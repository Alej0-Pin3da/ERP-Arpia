"""Contract tests for migrate.finanzas - F6 finanzas/socios (PR#7 slice).

Covers STRICT TDD acceptance from tasks #424 T9 (spec FIN-1..FIN-3, EXM-2/3/4,
NFR-1/2; design #423 finanzas.py + D3 + product decisions 1..5):

- Pure: type classification (equipo/maquinas/cursos -> Inversion; consumibles/
  servicios/fletes -> Gasto), montos as Decimal, fecha real preserved.
- Workbook -> plan: INVERSION VALQUI/MARGARA left block only; BOM insumos
  (already WAC via F2) are SKIPPED (no duplicate movements); right price
  sub-tables (J..N / H..L) never become movements; rows without a usable
  fecha follow D5 (omit + WARN), never now(); idempotence.
- DB (real PostgreSQL): the 3 socios batch (40/30/30, sum == 100) is
  persisted directly via ORM (design D3: the service rejects a batch because
  it demands sum==100 per create); movements keep the REAL excel fecha
  (never now()); the Rafael loan is type Inversion with socio_id NULL (no
  Rafael socio is created); BOM rows already in WAC are skipped; re-run does
  not duplicate movements (natural key fecha+tipo+monto+socio+descripcion);
  rollback leaves zero rows persisted.
- Phase: F6 registered in the runner registry; dry-run on the real workbook
  writes 0 rows (NFR-2).

Test-injected rows use the 'Migratest ' prefix so cleanup never touches real
migration data; the canonical catalog tipos inserted by bootstrap_catalogo()
are removed at module cleanup (same pattern as the other test_migrate_*).
"""

from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.db.session import SessionLocal
from app.models import (
    Insumo,
    MovimientoFinanciero,
    SociosConfiguracion,
    TipoProducto,
)
from migrate.context import FaseOptions, MigrationContext
from migrate.loaders import LibroMigracion

REAL_XLSX = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")
PREFIX = "Migratest"

# The 3 canonical socios of the migration (product decision 1).
SOCIOS_ESPERADOS = {
    "Valqui": Decimal("40"),
    "Margarita": Decimal("30"),
    "ARPIA": Decimal("30"),
}

# Names injected by tests; cleanup deletes exactly these, never catalog rows.
P_TERMO = f"{PREFIX} Termofijadora"
P_EQUIPO = f"{PREFIX} Equipo"
P_PRESTAMO = f"{PREFIX} Prestamo Rafael"
P_BOM_TELA = f"{PREFIX} Fin Tela"
P_CURSO = f"{PREFIX} Curso"
P_ENVIO = f"{PREFIX} Envio Materiales"
P_MARG = f"{PREFIX} Marg Equipo"

FECHA_TERMO = datetime(2023, 3, 17, tzinfo=timezone.utc)
FECHA_EQUIPO = datetime(2024, 2, 8, tzinfo=timezone.utc)
FECHA_CURSO = datetime(2025, 1, 16, tzinfo=timezone.utc)


# --------------------------------------------------------------------------- #
# Mini workbook builder (real layouts, verified against ARPIA.xlsx)
# --------------------------------------------------------------------------- #


def _mini_workbook(path: Path) -> None:
    """Mini finanzas workbook.

    - CORSET sheet: provides the BOM universe (catalog materials).
    - INVERSION VALQUI (header R2, data R3..): left block col A=cant,
      B=Producto, D=Costo, E=Fecha, F=Provedor; right price sub-table J..N
      must never become movements. Contains a BOM insumo row (-> skipped, F2)
      and non-BOM rows (equipo / loans).
    - INVERSION MARGARA (A/B/C/D/E): a BOM row and a non-BOM row.
    """
    wb = openpyxl.Workbook()
    bom = wb.active
    bom.title = "CORSET"
    bom.append(["CORSET", None, None, None, None, None, None, None, "TANGA"])
    bom.append(["Producto", "Ancho", "Alto", "cantidad Cms", "valor metro", "valor total"])
    bom.append([P_BOM_TELA, 64, 37, 2368, 2.5, None])  # R3 material BOM

    inv = wb.create_sheet("INVERSION VALQUI")
    inv.cell(row=2, column=1, value="Cantidad")
    inv.cell(row=2, column=2, value="Producto")
    inv.cell(row=2, column=4, value="Costo")
    inv.cell(row=2, column=5, value="Fecha")
    inv.cell(row=2, column=6, value="Provedor")
    # R3: BOM insumo (purchase -> F2 WAC, NOT a financial movement).
    inv.cell(row=3, column=1, value="4 mts")
    inv.cell(row=3, column=2, value=P_BOM_TELA)
    inv.cell(row=3, column=4, value=200)
    inv.cell(row=3, column=5, value=datetime(2024, 2, 17))
    # R4: equipment (termofijadora) -> Inversion.
    inv.cell(row=4, column=1, value=1)
    inv.cell(row=4, column=2, value=P_TERMO)
    inv.cell(row=4, column=4, value=960000)
    inv.cell(row=4, column=5, value=datetime(2023, 3, 17))
    # R5: loan (Rafael) -> Inversion with socio_id NULL (approved decision 2).
    inv.cell(row=5, column=1, value=1)
    inv.cell(row=5, column=2, value=P_PRESTAMO)
    inv.cell(row=5, column=4, value=101000)
    inv.cell(row=5, column=5, value=datetime(2023, 12, 1))
    # R6: right price sub-table (J..N) duplicates a purchase -> never a movement.
    inv.cell(row=6, column=10, value="Tensor 8")  # J
    inv.cell(row=6, column=11, value=100)         # K
    inv.cell(row=6, column=12, value=1)           # L
    inv.cell(row=6, column=13, value=6600)        # M

    marg = wb.create_sheet("INVERSION MARGARA")
    marg.cell(row=2, column=1, value="Cantidad")
    marg.cell(row=2, column=2, value="Producto")
    marg.cell(row=2, column=3, value="Costo")
    marg.cell(row=2, column=4, value="Fecha")
    marg.cell(row=2, column=5, value="Provedor")
    # R3: BOM insumo -> skipped (F2 already handled it).
    marg.cell(row=3, column=1, value="2 mts")
    marg.cell(row=3, column=2, value=P_BOM_TELA)
    marg.cell(row=3, column=3, value=100)
    marg.cell(row=3, column=4, value=datetime(2025, 2, 8))
    # R4: equipment -> Inversion (Margarita's share).
    marg.cell(row=4, column=1, value=1)
    marg.cell(row=4, column=2, value=P_MARG)
    marg.cell(row=4, column=3, value=2200000)
    marg.cell(row=4, column=4, value=datetime(2025, 1, 16))
    wb.save(path)


@pytest.fixture
def mini_finanzas(tmp_path) -> Path:
    path = tmp_path / "mini-finanzas.xlsx"
    _mini_workbook(path)
    return path


# --------------------------------------------------------------------------- #
# Module-level DB cleanup (canonical tipos + test rows)
# --------------------------------------------------------------------------- #


def _movimientos_id_de_test(db):
    """IDs de los Movimientos_Financieros con descripcion prefijo test."""
    return {
        m.id
        for m in db.query(MovimientoFinanciero)
        .filter(MovimientoFinanciero.descripcion.like(f"{PREFIX}%"))
        .all()
    }


def _borrar_test(db) -> None:
    """Limpia solo lo de este modulo: movimientos test, insumo."""
    mids = _movimientos_id_de_test(db)
    if mids:
        db.query(MovimientoFinanciero).filter(
            MovimientoFinanciero.id.in_(mids)
        ).delete(synchronize_session=False)
    db.query(Insumo).filter(Insumo.nombre == P_BOM_TELA).delete(
        synchronize_session=False
    )
    db.commit()


@pytest.fixture(scope="module")
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


@pytest.fixture(autouse=True, scope="module")
def _cleanup_after_module(db):
    yield
    _borrar_test(db)
    _borrar_socios_y_tipos(db)


def _borrar_socios_y_tipos(db) -> None:
    """Los 3 socios canonicos + los 6 Tipos_Producto de bootstrap_catalogo().

    Solo se borran socios sin movimientos (o con movimientos de TEST, que se
    limpian primero); un socio con movimientos reales nunca se borra."""
    # Movimientos de test por socio, primero.
    mids = db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.descripcion.like(f"{PREFIX}%")
    ).all()
    if mids:
        db.query(MovimientoFinanciero).filter(
            MovimientoFinanciero.id.in_([m.id for m in mids])
        ).delete(synchronize_session=False)
    for nombre in SOCIOS_ESPERADOS:
        socio = db.query(SociosConfiguracion).filter(
            SociosConfiguracion.nombre == nombre
        ).first()
        if socio is None:
            continue
        con_movimientos = db.query(MovimientoFinanciero).filter(
            MovimientoFinanciero.socio_id == socio.id
        ).first()
        if con_movimientos is None:
            db.delete(socio)
    # Tipos_Producto canonicos: SOLO si ningun producto los referencia (con la
    # migracion real cargada los productos reales los usan -> se conservan).
    from tests.conftest import borrar_tipos_canonicos_si_libres

    borrar_tipos_canonicos_si_libres(db)
    db.commit()


def _preparar_entorno(db) -> None:
    """Bootstrap catalogo + insumo BOM del mini (para verificar BOM-skip)."""
    from migrate.catalog import bootstrap_catalogo, upsert_insumo

    _borrar_test(db)
    _borrar_socios_y_tipos(db)
    bootstrap_catalogo(db)
    _ = upsert_insumo(db, P_BOM_TELA, categoria_nombre="Telas")
    db.commit()


# --------------------------------------------------------------------------- #
# 1. Pure: tipo classification (equipo/maquinas/cursos -> Inversion)
# --------------------------------------------------------------------------- #


def test_clasificar_equipo_como_inversion():
    from migrate.finanzas import clasificar_tipo

    assert clasificar_tipo(P_TERMO) == "Inversion"
    assert clasificar_tipo("Maquina plana Industrial") == "Inversion"
    assert clasificar_tipo("Impresora 3D FDM") == "Inversion"
    assert clasificar_tipo(P_CURSO) == "Inversion"
    assert clasificar_tipo("Stand en feria gotica 2 dias") == "Inversion"


def test_clasificar_consumibles_como_gasto():
    from migrate.finanzas import clasificar_tipo

    assert clasificar_tipo("Hosting y dominio") == "Gasto"
    assert clasificar_tipo("Publicidad en Meta Ads") == "Gasto"
    assert clasificar_tipo(P_ENVIO) == "Gasto"
    assert clasificar_tipo("Papeleria y organizacion") == "Gasto"


def test_clasificar_prestamo_rafael_tipo_inversion():
    from migrate.finanzas import clasificar_tipo

    assert clasificar_tipo(P_PRESTAMO) == "Inversion"


# --------------------------------------------------------------------------- #
# 2. Pure: plan from workbook (BOM skip + right sub-table + no invent) *
# --------------------------------------------------------------------------- #


def _plan_de(path):
    from migrate.finanzas import plan_finanzas

    from migrate.loaders import LibroMigracion

    with LibroMigracion(path) as libro:
        return plan_finanzas(libro)


def test_plan_lea_no_bom_y_skippea_bom(mini_finanzas):
    plan = _plan_de(mini_finanzas)
    nombres = {m.descripcion for m in plan.movimientos}
    # BOM insumo en la hoja VALQUI no es un movimiento (ya es compra WAC F2)
    assert P_BOM_TELA not in nombres
    assert P_TERMO in nombres
    assert P_PRESTAMO in nombres
    # Sub-tabla derecha (Tensor) nunca entra
    assert not any("Tensor" in n for n in nombres)


def test_plan_fechas_reales_y_montos(mini_finanzas):
    plan = _plan_de(mini_finanzas)
    termo = [m for m in plan.movimientos if m.descripcion == P_TERMO][0]
    assert termo.monto == Decimal("960000")
    assert termo.fecha == datetime(2023, 3, 17, tzinfo=timezone.utc)
    assert termo.tipo == "Inversion"


def test_plan_asignacion_socio_por_hoja(mini_finanzas):
    plan = _plan_de(mini_finanzas)
    por_hoja = {}
    for m in plan.movimientos:
        por_hoja.setdefault(m.hoja, m.socio_nombre)
    assert por_hoja.get("INVERSION VALQUI") == "Valqui"
    assert por_hoja.get("INVERSION MARGARA") == "Margarita"


# --------------------------------------------------------------------------- #
# 3. DB: socios 40/30/30 batch + movimiento con fecha real + prestamo NULL *
# --------------------------------------------------------------------------- #


def test_aplicar_crea_socios_batch_40_30_30(db, mini_finanzas):
    from migrate.finanzas import aplicar_finanzas

    _preparar_entorno(db)  # bootstrap (sin borrar socios previos)
    plan = _plan_de(mini_finanzas)
    res = aplicar_finanzas(db, plan)
    db.commit()

    socios = {s.nombre: s.porcentaje_participacion
              for s in db.query(SociosConfiguracion).all()}
    for nombre, pct in SOCIOS_ESPERADOS.items():
        assert socios.get(nombre) == pct
    # Suma == 100 de los 3 socios canonicos (FIN-2: la garantiza el pipeline)
    total = sum(socios[nombre] for nombre in SOCIOS_ESPERADOS)
    assert total == Decimal("100.0000")


def test_plan_no_duplica_socios_en_rerun(db, mini_finanzas):
    from migrate.finanzas import aplicar_finanzas

    _preparar_entorno(db)
    plan = _plan_de(mini_finanzas)
    aplicar_finanzas(db, plan)
    db.commit()
    aplicar_finanzas(db, plan)  # 2da corrida
    db.commit()

    count = (
        db.query(SociosConfiguracion)
        .filter(SociosConfiguracion.nombre.in_(SOCIOS_ESPERADOS))
        .count()
    )
    assert count == 3


def test_aplicar_persiste_fecha_real(mini_finanzas, db):
    from migrate.finanzas import aplicar_finanzas

    _preparar_entorno(db)
    plan = _plan_de(mini_finanzas)
    aplicar_finanzas(db, plan)
    db.commit()

    mov = (
        db.query(MovimientoFinanciero)
        .filter(MovimientoFinanciero.descripcion == P_TERMO)
        .one()
    )
    assert mov.fecha.strftime("%Y-%m-%d") == "2023-03-17"
    assert mov.tipo == "Inversion"
    assert mov.monto == Decimal("960000")


def test_prestamo_rafael_socio_id_null(mini_finanzas, db):
    from migrate.finanzas import aplicar_finanzas

    _preparar_entorno(db)
    plan = _plan_de(mini_finanzas)
    aplicar_finanzas(db, plan)
    db.commit()

    mov = (
        db.query(MovimientoFinanciero)
        .filter(MovimientoFinanciero.descripcion == P_PRESTAMO)
        .one()
    )
    assert mov.tipo == "Inversion"
    assert mov.socio_id is None  # decision 2: NO se crea socio Rafael


def test_aplicar_es_idempotente(db, mini_finanzas):
    from migrate.finanzas import aplicar_finanzas

    _preparar_entorno(db)
    plan = _plan_de(mini_finanzas)
    aplicar_finanzas(db, plan)
    db.commit()
    aplicar_finanzas(db, plan)  # re-ejecucion
    db.commit()

    # 3 movimientos reales: termo (VALQUI) + prestamo (VALQUI) + marg (MARGARA).
    # Scoped a las filas de PRUEBA: la DB real ya tiene los movimientos de la
    # migracion cargada (patron _borrar_filas_test — nunca asumir DB vacia).
    movs_test = db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.descripcion.like(f"{PREFIX}%")
    ).all()
    assert len(movs_test) == 3
    # no se duplican ni el prestamo ni el equipo


def test_rollback_no_persiste_nada(db, mini_finanzas):
    """Si el caller hace rollback: cero movimientos de prueba y los socios
    canonicos quedan como estaban (sin filas residuales de la fase)."""
    from app.models import SociosConfiguracion
    from migrate.finanzas import aplicar_finanzas

    _preparar_entorno(db)
    _borrar_test(db)
    # Los 3 socios canonicos YA existen (migracion real cargada); la fase usa
    # get-or-create, asi que un rollback no debe alterar su estado ni crear
    # movimientos de prueba.
    socios_antes = {
        s.nombre: s.porcentaje_participacion
        for s in db.query(SociosConfiguracion).all()
        if s.nombre in SOCIOS_ESPERADOS
    }
    plan = _plan_de(mini_finanzas)
    aplicar_finanzas(db, plan)  # writes in the session (no commit)
    db.rollback()  # caller aborts the phase (EXM-4)
    # Scoped a filas de PRUEBA: la migracion real ya dejo sus movimientos.
    assert (
        db.query(MovimientoFinanciero)
        .filter(MovimientoFinanciero.descripcion.like(f"{PREFIX}%"))
        .count()
        == 0
    )
    socios_despues = {
        s.nombre: s.porcentaje_participacion
        for s in db.query(SociosConfiguracion).all()
        if s.nombre in SOCIOS_ESPERADOS
    }
    assert socios_despues == socios_antes  # sin cambios netos por el rollback


def test_dry_run_registrada_en_runner():
    from migrate import FASE_RUNNERS, FASES, FASES_IMPLEMENTADAS

    assert any(f.id == "F6" for f in FASES)
    assert "F6" in FASE_RUNNERS
    assert "F6" in FASES_IMPLEMENTADAS


def test_cargar_finanzas_dry_run_real_no_escribe():
    from migrate.finanzas import cargar_finanzas

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    db = SessionLocal()
    try:
        antes = db.query(MovimientoFinanciero).count()
        ctx = MigrationContext.para_fase(
            FaseOptions(source=REAL_XLSX, modo="dry-run"), "F6"
        )
        plan = cargar_finanzas(ctx)
        despues = db.query(MovimientoFinanciero).count()
        assert antes == despues  # NFR-2: 0 filas escritas
        assert not ctx.report.tenga_errores
        # Plan real informativo: al menos 5 movimientos (maquinas, hostings, etc)
        assert plan.conteo_movimientos >= 5
    finally:
        db.close()
"""Contract tests for migrate.purchases - F2 historical WAC purchases (PR#4).

Covers the STRICT TDD acceptance from tasks #424 T5 (spec EXM-2/3/4/5, D1/D5,
NFR-1/2):

- Pure: quantity normalization to the insumo's canonical unit ('4 mts'->4 m,
  '50 cm'->0.5 m, numeric->as-is), non-inferable values rejected.
- Workbook -> plan: ONLY catalog (BOM) insumos become WAC purchases; non-BOM
  rows (equipo, gastos) are excluded (-> finanzas F6); right-hand price
  sub-tables (J..N / H..L) that duplicate left blocks are never loaded; empty
  células de fecha follow D5 (inherit contiguous same insumo, otherwise omit +
  WARN), never now().
- DB (real PostgreSQL): applying the plan registers CompraInsumo rows with the
  REAL excel fecha, updates insumo stock_actual + costo_promedio_actual via
  wac.registrar_compra (WAC formula), is idempotent on re-run (natural key
  insumo+fecha+cantidad+precio), and a rollback leaves cero filas persistidas
  (no internal commit of the service).
- Phase: F2 registered in the runner registry; dry-run on the real workbook
  writes 0 rows (NFR-2) and reports the BOM purchase plan.

Test-injected rows use the 'MigraTest ' prefix so cleanup never touches real
migration data.
"""

from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.db.session import SessionLocal
from app.models import CompraInsumo, Insumo
from migrate.context import FaseOptions, MigrationContext
from migrate.loaders import LibroMigracion

REAL_XLSX = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")
PREFIX_TEST = "Migratest"

DIA = datetime(2026, 2, 17, tzinfo=UTC)
DIA2 = datetime(2026, 3, 3, tzinfo=UTC)


# --------------------------------------------------------------------------- #
# Fixtures / helpers
# --------------------------------------------------------------------------- #


def _mini_workbook(path: Path) -> None:
    """Mini workbook: a CORSET sheet providing BOM materials (catalog universe)
    + INVERSION VALQUI with the Excel purchase layout (left block A..E, right
    price sub-table J..N starting at R13)."""
    wb = openpyxl.Workbook()
    bom = wb.active
    bom.title = "CORSET"
    bom.append(["CORSET", None, None, None, None, None, None, None, "TANGA"])
    bom.append(["Producto", "Ancho", "Alto", "cantidad Cms", "valor metro", "valor total"])
    bom.append([f"{PREFIX_TEST} Tela", 64, 37, 2368, 2.5, None])  # R3 material BOM
    bom.append([f"{PREFIX_TEST} Argolla", 2, 1, 2, 72, None])
    bom.append([f"{PREFIX_TEST} Encaje", 28, 18, 504, 4.0, None])
    bom.append(["Horas trabajo", None, None, None, None, None])
    bom.append(["COSTO TOTAL CONJUNTO", None, None, None, None, None])

    inv = wb.create_sheet("INVERSION VALQUI")
    inv.cell(row=2, column=1, value="Cantidad")
    inv.cell(row=2, column=2, value="Producto")
    inv.cell(row=2, column=4, value="Costo")
    inv.cell(row=2, column=5, value="Fecha")
    inv.cell(row=2, column=6, value="Provedor")
    # R3: fabric purchase 4 mts @ 200 total -> unit price 50
    inv.cell(row=3, column=1, value="4 mts")
    inv.cell(row=3, column=2, value=f"{PREFIX_TEST} Tela")
    inv.cell(row=3, column=4, value=200)
    inv.cell(row=3, column=5, value=datetime(2026, 2, 17))
    # R4: another BOM purchase with empty fecha (R6 hereda de R3 mismo insumo? no, distinto)
    inv.cell(row=4, column=1, value=12)
    inv.cell(row=4, column=2, value=f"{PREFIX_TEST} Argolla")
    inv.cell(row=4, column=4, value=2400)
    inv.cell(row=4, column=5, value=datetime(2026, 3, 3))
    # R5: EQUIPO row (no BOM) -> excluded
    inv.cell(row=5, column=1, value=1)
    inv.cell(row=5, column=2, value=f"{PREFIX_TEST} Equipo")
    inv.cell(row=5, column=4, value=5000)
    inv.cell(row=5, column=5, value=datetime(2026, 3, 3))
    # R6: BOM compra SIN fecha -> hereda DIA2? no: mismo insumo sin fecha previa -> omitted+WARN
    inv.cell(row=6, column=1, value="5 mts")
    inv.cell(row=6, column=2, value=f"{PREFIX_TEST} Encaje")
    inv.cell(row=6, column=4, value=1200)
    # R7: BOM compra SIN fecha pero mismo insumo (Tela) que R3 -> hereda DIA
    inv.cell(row=7, column=1, value="2 mts")
    inv.cell(row=7, column=2, value=f"{PREFIX_TEST} Tela")
    inv.cell(row=7, column=4, value=60)
    # Derecha: la sub-tabla de precios duplica la compra (400cm x 1cm = 4 mts a M=200)
    inv.cell(row=14, column=10, value=f"{PREFIX_TEST} Tela")  # J
    inv.cell(row=14, column=11, value=400)  # K Largo CMS
    inv.cell(row=14, column=12, value=1)  # L Ancho CMS
    inv.cell(row=14, column=13, value=200)  # M Valor
    wb.save(path)


@pytest.fixture
def mini_libro(tmp_path) -> Path:
    path = tmp_path / "mini-purchases.xlsx"
    _mini_workbook(path)
    return path


def _plan_de(mini_libro):
    from migrate.purchases import plan_compras

    with LibroMigracion(mini_libro) as libro:
        return plan_compras(libro)


def _preparar_catalogo(db) -> dict[str, int]:
    """F1 bootstrap + insumos BOM del mini: insumos de prueba."""
    from migrate.catalog import bootstrap_catalogo, upsert_insumo

    bootstrap_catalogo(db)
    db.flush()
    ids = {}
    ids[PREFIX_TEST + " Tela"] = upsert_insumo(
        db, f"{PREFIX_TEST} Tela", categoria_nombre="Telas"
    ).id
    ids[PREFIX_TEST + " Argolla"] = upsert_insumo(
        db, f"{PREFIX_TEST} Argolla", categoria_nombre="Herrajes"
    ).id
    ids[PREFIX_TEST + " Encaje"] = upsert_insumo(
        db, f"{PREFIX_TEST} Encaje", categoria_nombre="Telas"
    ).id
    db.commit()
    return ids


def _borrar_filas_test(db) -> None:
    """Remove rows injected by this test module (exact-name matches only)."""
    from app.models import CompraInsumo, Insumo, TipoProducto

    nombres_insumo = [f"{PREFIX_TEST} Tela", f"{PREFIX_TEST} Argolla", f"{PREFIX_TEST} Encaje"]
    insumos = db.query(Insumo).filter(Insumo.nombre.in_(nombres_insumo)).all()
    for insumo in insumos:
        db.query(CompraInsumo).filter(CompraInsumo.insumo_id == insumo.id).delete(
            synchronize_session=False
        )
    db.query(Insumo).filter(Insumo.nombre.in_(nombres_insumo)).delete(synchronize_session=False)
    # Remove the canonical catalog tipos that bootstrap_catalogo() inserts.
    # They are migration content, not app seed data; leaving them pollutes the
    # shared per-sheet DB and breaks pagination tests that assume an empty table.
    db.query(TipoProducto).filter(
        TipoProducto.nombre.in_(["Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"])
    ).delete(synchronize_session=False)
    db.commit()


@pytest.fixture(autouse=True, scope="module")
def _cleanup_after_module():
    yield
    db = SessionLocal()
    try:
        _borrar_filas_test(db)
    finally:
        db.close()


@pytest.fixture(scope="module")
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


# --------------------------------------------------------------------------- #
# 1. Pure: quantity normalization (canonical unit)
# --------------------------------------------------------------------------- #


def test_normalizar_cantidad_compra_metros_y_numerico():
    from migrate.purchases import normalizar_cantidad_compra

    assert normalizar_cantidad_compra("4 mts", "m") == 4
    assert normalizar_cantidad_compra("2,90 mts", "m") == Decimal("2.9")
    assert normalizar_cantidad_compra(12, "un") == 12
    assert normalizar_cantidad_compra(0.5, "m") == 0.5


def test_normalizar_cantidad_compra_cm_a_metros():
    from migrate.purchases import normalizar_cantidad_compra

    assert normalizar_cantidad_compra("50 cm", "m") == 0.5
    assert normalizar_cantidad_compra("100 cm", "m") == 1


def test_normalizar_cantidad_compra_rechaza_no_interpretables():
    from migrate.purchases import normalizar_cantidad_compra

    assert normalizar_cantidad_compra(None, "m") is None
    assert normalizar_cantidad_compra("#DIV/0!", "m") is None
    assert normalizar_cantidad_compra("un metro aprox", "m") is None


def test_normalizar_cantidad_compra_expresion_area_a_m2():
    """EXM-2 borde: '50 x 280 cm' (expresion area en telas) -> 1.4 m2.

    Real case: INVERSION MARGARA A7 = '50 x 280 cm' for a Telas insumo.
    The area expression is interpreted as an area, not a length, and only for
    the 'm' canonical target (design D4: area vs largo segun el insumo).
    """
    from migrate.purchases import normalizar_cantidad_compra

    assert normalizar_cantidad_compra("50 x 280 cm", "m") == Decimal("1.4")
    # Not an area expression: length strings keep their current behavior.
    assert normalizar_cantidad_compra("10 mts", "m") == Decimal("10")
    # Area is Telas-only: 'un' target must not consume the expression.
    assert normalizar_cantidad_compra("50 x 280 cm", "un") is None


# --------------------------------------------------------------------------- #
# 2. Workbook -> plan: BOM filter, right sub-table, fecha policy
# --------------------------------------------------------------------------- #


def test_plan_compras_solo_insumos_bom(mini_libro):
    plan = _plan_de(mini_libro)
    nombres = {c.insumo_nombre for c in plan.compras}
    assert f"{PREFIX_TEST} Tela" in nombres
    assert f"{PREFIX_TEST} Argolla" in nombres
    # R5 es EQUIPO (no esta en el universos BOM): no es compra WAC
    assert not any(c.insumo_nombre == f"{PREFIX_TEST} Equipo" for c in plan.compras)
    # La sub-tabla derecha (R14) no genera compra duplicada
    assert plan.conteo_compras == 3


def test_plan_compras_calcula_precio_unitario(mini_libro):
    plan = _plan_de(mini_libro)
    # First-wins: R3 (4 mts @ 50) y R7 (2 mts @ 30) son dos compras del mismo
    # insumo; la heredada no debe pisar la fila original en el lookup.
    por = {}
    for c in plan.compras:
        por.setdefault(c.insumo_nombre, c)
    assert por[f"{PREFIX_TEST} Tela"].cantidad == Decimal(4)
    assert por[f"{PREFIX_TEST} Tela"].precio_unitario == Decimal(50)
    assert por[f"{PREFIX_TEST} Argolla"].cantidad == 12
    assert por[f"{PREFIX_TEST} Argolla"].precio_unitario == Decimal(200)


def test_plan_compras_fecha_heredada_contigua(mini_libro):
    plan = _plan_de(mini_libro)
    # R7 hereda la fecha DIA de la compra R3 (mismo insumo Tela)
    tela_2m = [c for c in plan.compras if c.insumo_nombre == f"{PREFIX_TEST} Tela"]
    heredada = [c for c in tela_2m if c.fecha_heredada]
    assert len(heredada) == 1
    # (fecha no normalizada -> coerce a aware utc)
    assert heredada[0].fecha == DIA


def test_plan_compras_sin_fecha_ni_heredable_omitida(mini_libro):
    plan = _plan_de(mini_libro)
    # R6 (Encaje sin fecha y sin previo del mismo insumo) no debe entrar
    assert plan.conteo_compras == 3
    assert not any(c.insumo_nombre == f"{PREFIX_TEST} Encaje" for c in plan.compras)
    assert plan.conteos.sin_fecha == 1


# --------------------------------------------------------------------------- #
# 2b. P1 fix: sub-tabla derecha como fuente UNICA (no duplicada)
# --------------------------------------------------------------------------- #


def _mini_workbook_derecha(path: Path) -> None:
    """Mini con sub-tabla derecha REAL (layout VALQUI J..N).

    - CORSET: 4 materiales BOM (Tela, Argolla, + dos que solo existen en la
      derecha: 'Tela Derecha' y 'Argolla Derecha').
    - VALQUI: left block R3..R5 (compra Tela, Argolla, Tela Derecha) y la
      sub-tabla derecha J..N:
      * R15 J='Argolla Derecha' (K=100, L=1, M=7200) -> fuente UNICA: la
        derecha no duplica la izquierda -> compra 100 un @ 72, fecha E15.
      * R16 J='Tela Derecha' (K=400, L=1, M=200) -> DUPLICA la compra
        izquierda R5 (mismo item) -> se descarta como antes.
    """
    wb = openpyxl.Workbook()
    bom = wb.active
    bom.title = "CORSET"
    bom.append(["CORSET", None, None, None, None, None, None, None, "TANGA"])
    bom.append(["Producto", "Ancho", "Alto", "cantidad Cms", "valor metro", "valor total"])
    bom.append([f"{PREFIX_TEST} Tela", 64, 37, 2368, 2.5, None])
    bom.append([f"{PREFIX_TEST} Argolla", 2, 1, 2, 72, None])
    bom.append([f"{PREFIX_TEST} Tela Derecha", 64, 37, 2368, 2.5, None])
    bom.append([f"{PREFIX_TEST} Argolla Derecha", 2, 1, 2, 72, None])

    inv = wb.create_sheet("INVERSION VALQUI")
    inv.cell(row=2, column=1, value="Cantidad")
    inv.cell(row=2, column=2, value="Producto")
    inv.cell(row=2, column=4, value="Costo")
    inv.cell(row=2, column=5, value="Fecha")
    inv.cell(row=2, column=6, value="Provedor")
    inv.cell(row=13, column=10, value="Producto")  # header sub-tabla J..N
    inv.cell(row=13, column=11, value="Largo CMS")
    inv.cell(row=13, column=12, value="Ancho CMS")
    inv.cell(row=13, column=13, value="Valor")
    inv.cell(row=13, column=14, value="Unitario")
    # R3..R5: bloque izquierdo
    inv.cell(row=3, column=1, value="4 mts")
    inv.cell(row=3, column=2, value=f"{PREFIX_TEST} Tela")
    inv.cell(row=3, column=4, value=200)
    inv.cell(row=3, column=5, value=datetime(2026, 2, 17))
    inv.cell(row=4, column=1, value=12)
    inv.cell(row=4, column=2, value=f"{PREFIX_TEST} Argolla")
    inv.cell(row=4, column=4, value=2400)
    inv.cell(row=4, column=5, value=datetime(2026, 2, 17))
    inv.cell(row=5, column=1, value="4 mts")
    inv.cell(row=5, column=2, value=f"{PREFIX_TEST} Tela Derecha")
    inv.cell(row=5, column=4, value=200)
    inv.cell(row=5, column=5, value=datetime(2026, 2, 17))
    # R15: derecha fuente unica -> Argolla Derecha 100 un @ 72 (fecha E15)
    inv.cell(row=15, column=10, value=f"{PREFIX_TEST} Argolla Derecha")
    inv.cell(row=15, column=11, value=100)
    inv.cell(row=15, column=12, value=1)
    inv.cell(row=15, column=13, value=7200)
    inv.cell(row=15, column=5, value=datetime(2026, 2, 17))
    # R16: derecha duplica izquierda (Tela Derecha ya comprada en R5 con la
    # MISMA fecha de operacion) -> descarte
    inv.cell(row=16, column=10, value=f"{PREFIX_TEST} Tela Derecha")
    inv.cell(row=16, column=11, value=400)
    inv.cell(row=16, column=12, value=1)
    inv.cell(row=16, column=13, value=200)
    inv.cell(row=16, column=5, value=datetime(2026, 2, 17))
    wb.save(path)


@pytest.fixture
def mini_libro_derecha(tmp_path) -> Path:
    path = tmp_path / "mini-purchases-derecha.xlsx"
    _mini_workbook_derecha(path)
    return path


def test_plan_compras_derecha_fuente_unica_genera_compra(mini_libro_derecha):
    """P1 fix: un item de la sub-tabla derecha NO presente en el bloque
    izquierdo de la misma hoja es fuente UNICA -> se procesa como compra
    (cantidad = K en la unidad canonica, precio = M / K, fecha del bloque
    izquierdo de la misma fila)."""
    plan = _plan_de(mini_libro_derecha)
    compra = next(c for c in plan.compras if c.insumo_nombre == f"{PREFIX_TEST} Argolla Derecha")
    assert compra.cantidad == 100  # K=100 en 'un' (Herrajes)
    assert compra.precio_unitario == 72  # M/K = 7200/100
    assert compra.fecha == DIA  # fecha del left E de la misma fila
    assert compra.hoja == "INVERSION VALQUI"


def test_plan_compras_derecha_misma_fecha_duplica_se_descarta(mini_libro_derecha):
    """Regression (2026-08): la sub-tabla derecha duplica la izquierda SOLO
    cuando el item se compro a la izquierda con la MISMA fecha de operacion
    (mismo registro en dos bloques). Aqui R16 'Tela Derecha' no tiene fecha
    propia -> hereda la de la compra izquierda contigua (R5, DIA) -> es la
    MISMA compra -> se descarta como duplicado (el workbook pone el mismo
    item en ambos bloques de una misma operacion)."""
    plan = _plan_de(mini_libro_derecha)
    telas_derecha = [c for c in plan.compras if c.insumo_nombre == f"{PREFIX_TEST} Tela Derecha"]
    assert len(telas_derecha) == 1  # solo la compra izquierda R5, la derecha se descarta
    assert telas_derecha[0].fila == 5


def test_plan_compras_derecha_fecha_distinta_no_es_duplicado():
    """Regression (2026-08, workbook real): 'Cadena plateada gruesa totebag'
    aparece a la izquierda (F96, 2026-03-27, 1 mts) y en la sub-tabla derecha
    (F54, 2024-07-23, 100 mts). Son DOS compras legitimas en fechas distintas.
    El criterio anterior descartaba la derecha por nombre -> 35 compras de
    2024 se perdian y F5 fallaba con InsufficientStockError. La derecha solo
    duplica cuando la fecha coincide."""
    from migrate.purchases import plan_compras

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    with LibroMigracion(REAL_XLSX) as libro:
        plan = plan_compras(libro)

    cadenas = [c for c in plan.compras if "Cadena plateada gruesa" in c.insumo_nombre]
    fechas = sorted(str(c.fecha.date()) for c in cadenas)
    # La compra derecha F54 (2024) ya no se pierde.
    assert any(f.startswith("2024-") for f in fechas), f"cadena 2024 perdida: {fechas}"
    assert len(cadenas) >= 2


def test_plan_compras_derecha_duplicada_se_descarta(mini_libro_derecha):
    """P1 fix: cuando la derecha duplica un item YA comprado en la izquierda
    de la misma hoja, se descarta como antes (no se duplica WAC)."""
    plan = _plan_de(mini_libro_derecha)
    tela_derecha = [c for c in plan.compras if c.insumo_nombre == f"{PREFIX_TEST} Tela Derecha"]
    assert len(tela_derecha) == 1  # solo la compra izquierda R5
    assert tela_derecha[0].cantidad == 4  # 4 mts (no la copia de 400cm)
    assert plan.conteos.derecha >= 1  # la fila derecha duplicada se cuenta


# --------------------------------------------------------------------------- #
# 3. DB: aplicar + WAC + idempotencia + rollback
# --------------------------------------------------------------------------- #


def test_aplicar_plan_persiste_fechas_y_wac(db, mini_libro):
    from migrate.purchases import aplicar_compras

    _preparar_catalogo(db)
    plan = _plan_de(mini_libro)
    aplicar_compras(db, plan)
    db.commit()

    insumo_tela = db.query(Insumo).filter(Insumo.nombre == f"{PREFIX_TEST} Tela").one()
    compras = db.query(CompraInsumo).filter(CompraInsumo.insumo_id == insumo_tela.id).all()
    assert len(compras) == 2  # R3 + R7 heredada
    fechas = sorted(c.fecha_compra for c in compras)
    assert fechas[0] == DIA
    assert fechas[1] == DIA  # la heredada persiste la MISMA fecha real (no now())
    # WAC: (4 x 50 + 2 x 30) / 6 = 43.3333... (NUMERIC(15,4) storage)
    assert insumo_tela.stock_actual == 6
    assert insumo_tela.costo_promedio_actual == Decimal("43.3333")

    argolla = db.query(Insumo).filter(Insumo.nombre == f"{PREFIX_TEST} Argolla").one()
    compras_a = db.query(CompraInsumo).filter(CompraInsumo.insumo_id == argolla.id).all()
    assert len(compras_a) == 1
    assert argolla.stock_actual == 12
    assert argolla.costo_promedio_actual == 200


def test_aplicar_es_idempotente(db, mini_libro):
    from migrate.purchases import aplicar_compras

    ids = _preparar_catalogo(db)
    plan = _plan_de(mini_libro)
    aplicar_compras(db, plan)
    db.commit()
    aplicar_compras(db, plan)
    db.commit()

    from app.models import CompraInsumo

    tela = db.get(Insumo, ids[PREFIX_TEST + " Tela"])
    compras = db.query(CompraInsumo).filter(CompraInsumo.insumo_id == tela.id).all()
    assert len(compras) == 2  # sin duplicados tras re-ejecutaron
    assert tela.stock_actual == 6  # WAC no se re-aplica


def test_rollback_no_persiste_nada(db):
    """commit=False del servicio respetado: si el caller hace rollback, 0 filas."""
    from app.models import CompraInsumo
    from app.services.wac import registrar_compra

    ups = _preparar_catalogo(db)
    tela_id = ups[PREFIX_TEST + " Tela"]
    # Aislamiento: tests de plan previos ya insertaron compras de Tela.
    db.query(CompraInsumo).filter(CompraInsumo.insumo_id == tela_id).delete(
        synchronize_session=False
    )
    db.commit()
    registrar_compra(db, tela_id, 10, 25, fecha_compra=DIA, commit=False)
    db.rollback()
    assert db.query(CompraInsumo).filter(CompraInsumo.insumo_id == tela_id).count() == 0


# --------------------------------------------------------------------------- #
# 4. Phase: runner registry + real workbook dry run (NFR-2)
# --------------------------------------------------------------------------- #


def test_f2_registrada_en_runner():
    from migrate import FASE_RUNNERS, FASES, FASES_IMPLEMENTADAS

    assert any(f.id == "F2" for f in FASES)
    assert "F2" in FASE_RUNNERS
    assert "F2" in FASES_IMPLEMENTADAS


def test_cargar_compras_dry_run_real_no_escribe():
    from app.models import CompraInsumo
    from migrate.purchases import cargar_compras

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    db = SessionLocal()
    try:
        antes = db.query(CompraInsumo).count()
        ctx = MigrationContext.para_fase(FaseOptions(source=REAL_XLSX, modo="dry-run"), "F2")
        cargar_compras(ctx)
        despues = db.query(CompraInsumo).count()
        assert antes == despues  # NFR-2: 0 filas escritas
        assert not ctx.report.tenga_errores
        assert ctx.report.count("WARN") >= 7  # filas con fecha omitida/otros
    finally:
        db.close()

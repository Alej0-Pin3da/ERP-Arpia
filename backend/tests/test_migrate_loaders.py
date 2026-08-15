"""Tests for migrate.loaders - bounded per-sheet reading (PR#1 slice)."""

from pathlib import Path

import openpyxl
import pytest

from migrate.loaders import (
    SHEET_BOUNDS,
    HojaInexistenteError,
    LibroMigracion,
)


def _crear_libro(path: Path) -> None:
    """Mini ARPIA-like workbook: VENTAS + STICKERS inflated sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VENTAS"
    # Header row (R1) + 3 real rows (R2..R4) + junk M37 loose cell.
    ws.append(["Producto", "G", "H", "M", "P"])
    ws.append(["SET AELO", 80000, 38805, None, "celes"])
    ws.append(["TOTEBAG", 45000, 25765.09524, None, "Camila"])
    ws.append(["", 95000, 29826, None, "Olga"])
    ws["M37"] = 65618.01429  # loose junk cell (as in real VENTAS)
    # STICKERS: header + 4 real rows.
    st = wb.create_sheet("STICKERS")
    st.append(["TIPO", "URL", "Precio Unidad", "Ubicacion", "Contactado"])
    for nombre in ["Bexxhamel", "JM Confecciones", "SEHA Text", "ZureTex"]:
        st.append(["Camisetas", nombre, 11500, "Cali", "SI"])
    wb.save(path)


@pytest.fixture
def mini_libro(tmp_path) -> Path:
    path = tmp_path / "mini.xlsx"
    _crear_libro(path)
    return path


def test_lee_rango_acotado_ventas(mini_libro):
    with LibroMigracion(mini_libro) as libro:
        lect = libro.leer_hoja("VENTAS")
        # 3 filas reales (la 4ta fila del fixture no tiene producto, igual cuenta
        # como fila leida porque el loader acota por rango, no por semantica).
        assert len(lect.filas) >= 3


def test_rango_acotado_no_lee_fuera_del_rango(mini_libro):
    with LibroMigracion(mini_libro) as libro:
        lect = libro.leer_hoja("VENTAS")
        # Rango acotado 2..17: 3 filas reales; las vacias R5..R17 se descartan
        # (13); la celda suelta M37 queda fuera del rango y nunca es dato.
        assert len(lect.filas) == 3
        assert lect.descartadas == 13
        assert not any("M37" in fila for fila in lect.filas)


def test_stickers_matriz_acotada(mini_libro):
    with LibroMigracion(mini_libro) as libro:
        lect = libro.leer_hoja("STICKERS")
        # Bounds (2..33) acotan la lectura: 4 filas reales, sin vacias que descartar.
        assert len(lect.filas) == 4
        assert lect.descartadas == 0


def test_hoja_inexistente_error_claro(mini_libro):
    with LibroMigracion(mini_libro) as libro:
        with pytest.raises(HojaInexistenteError) as exc:
            libro.leer_hoja("NO EXISTE")
        assert "NO EXISTE" in str(exc.value)
        assert "mini.xlsx" in str(exc.value)


def test_m37_suelta_reportada(mini_libro):
    from migrate.report import Report

    report = Report(fase="test", modo="dry-run")
    with LibroMigracion(mini_libro) as libro:
        libro.leer_hoja("VENTAS", report=report)
    warns = [e for e in report.entradas if e.nivel == "WARN" and e.hoja == "VENTAS"]
    assert any(e.celda == "M37" for e in warns)


def test_camisetas_inv_aplica_mapeo_columnas_por_hoja(tmp_path):
    """EXM-1 error scenario: CAMISETAS INV header desalineado -> se aplica el
    mapeo de columnas propio de la hoja (D=tipo, E=origen, F=color reales)."""
    from migrate.loaders import COL_MAP_CAMISETAS_INV
    from migrate.report import Report

    path = tmp_path / "camisetas-inv.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CAMISETAS INV"
    # Header R1 desalineado: dice E=Fecha, F=Provedor pero los datos reales
    # ponen Tipo/Origen/Color en D/E/F (spec EXM-1 error scenario).
    ws.append(["Cantidad", "Talla", "Costo", None, "Fecha", "Provedor"])
    ws.append([1.0, "M", 11500.0, "Camiseta", "CALI", "Negra"])  # R2
    wb.save(path)

    report = Report(fase="test", modo="dry-run")
    with LibroMigracion(path) as libro:
        lect = libro.leer_hoja("CAMISETAS INV", report=report)
    assert len(lect.filas) == 1
    fila = lect.filas[0]
    # El mapeo por hoja traduce las columnas fisicas a semanticas.
    assert fila.get("Tipo") == "Camiseta"
    assert fila.get("Origen") == "CALI"
    assert fila.get("Color") == "Negra"
    assert fila.get("Cantidad") == 1.0
    # El encabezado desalineado se reporta (hoja/fila/celda, N7g).
    warns = [e for e in report.entradas if e.nivel == "WARN" and e.hoja == "CAMISETAS INV"]
    assert warns, "el encabezado desalineado debe reportarse como WARN"
    assert any("desalineado" in e.mensaje.lower() for e in warns)
    assert COL_MAP_CAMISETAS_INV["D"] == "Tipo"


def test_todas_las_hojas_registradas():
    # El libro real tiene 24 hojas; el spec de bounds debe cubrirlas todas.
    from openpyxl import load_workbook as lw

    real = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")
    if not real.exists():
        pytest.skip("ARPIA.xlsx no disponible")
    wb = lw(real, read_only=True, data_only=True)
    faltantes = [s for s in wb.sheetnames if s not in SHEET_BOUNDS]
    wb.close()
    assert faltantes == [], f"hojas sin bounds: {faltantes}"

"""Contract tests for migrate.sales - F5 ventas historicas (PR#6 slice).

Covers STRICT TDD acceptance from tasks #424 T8 (spec VTA-1..VTA-4, EXM-2/3/4,
NFR-1/2; design #423 sales.py + D2 + D7):

- Workbook -> plan: reads the VENTAS sheet within its bounded range; rows with
  no product in col A are SCOPE OUT (VTA-4, reported, never loaded); real
  columns: G=precio tal-cual, H=costo FULL, M=fecha real, P=cliente (D7),
  O=nota DESC (never re-applied, VTA-2).
- DB (real PostgreSQL): INSERT directo Venta + Detalle_Venta with fecha real,
  canal_venta='feria' (decision de producto; default del modelo), precio
  tal-cual (sin doble descuento), costo_unitario_aplicado = costo FULL del
  Excel (sin recalcular con el motor actual), cliente upsert por nombre.
- Destock: tras insertar todas las ventas se agrega la explosion BOM del
  servicio (no registrar_venta) y se descuenta stock en lote (FOR UPDATE,
  consistente con inventory.descontar_stock); stock insuficiente -> 409 ->
  rollback de fase (EXM-4), cero ventas residuales.
- Idempotencia (NFR-1/EXM-3): natural key (fecha, cliente, producto, variante,
  cantidad, precio, costo) con guard cuantitativo — re-ejecutar no duplica.
- Phase: F5 registrada en el runner; dry-run sobre ARPIA.xlsx real: 0
  escrituras (NFR-2).

Test-injected rows use the 'Migratest ' prefix so cleanup never touches real
migration data; the canonical catalog tipos inserted by bootstrap_catalogo()
are removed at module cleanup (same pattern as the other test_migrate_*).
"""

from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.core.exceptions import InsufficientStockError
from app.db.session import SessionLocal
from app.models import (
    BomInsumo,
    Cliente,
    DetalleVenta,
    Insumo,
    Producto,
    TipoProducto,
    VarianteProducto,
    Venta,
)
from migrate.context import FaseOptions, MigrationContext
from migrate.loaders import LibroMigracion

REAL_XLSX = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")

# Names injected by tests; cleanup deletes exactly these, never catalog rows.
P = "Migratest"
P_SET = f"{P} Venta Set"
P_TOTE = f"{P} Venta Tote"
P_CLI = f"{P} Venta Cliente"
P_CLI2 = f"{P} Venta Camila"
P_TELA = f"{P} Venta Tela"


# --------------------------------------------------------------------------- #
# Mini workbook builder (real VENTAS layout: header R1, data R2.., bounded 2..17)
# --------------------------------------------------------------------------- #


def _mini_workbook(path: Path) -> None:
    """Mini VENTAS: 2 real sales + 1 row without product (SCOPE OUT) + 1 junk
    totals row. Columns verified against ARPIA.xlsx (2026-08-08)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VENTAS"
    # Header (R1) — informational only (loader starts at R2).
    ws.append(
        [
            "Producto",
            "Set Ocipete",
            "Set Aelo",
            None,
            "Blusa Arpia ML",
            None,
            "Precio Venta",
            "Costo",
            "Ganancias",
            None,
            None,
            None,
            "Fecha",
            "Col 1",
            "Col 2",
            "Col 3",
        ]
    )
    # R2: product with variant + cliente NOTES (price already discounted).
    ws.append(
        [
            P_SET,
            None,
            "S",
            None,
            None,
            None,
            71250.0,
            26109,
            45141,
            None,
            None,
            None,
            datetime(2026, 3, 20),
            "vino",
            "DESC 25%",
            P_CLI,
        ]
    )
    # R3: product without variant/talla.
    ws.append(
        [
            P_TOTE,
            None,
            None,
            None,
            None,
            None,
            45000.0,
            25765.09524,
            19234.90476,
            None,
            None,
            None,
            datetime(2026, 4, 24),
            None,
            None,
            P_CLI2,
        ]
    )
    # R4: no product -> SCOPE OUT (VTA-4).
    ws.append(
        [
            None,
            None,
            None,
            None,
            None,
            None,
            95000.0,
            29826.0,
            65174,
            None,
            None,
            None,
            datetime(2026, 5, 9),
            None,
            None,
            "Olga",
        ]
    )
    # R5: junk totals (zeros) -> excluded too.
    ws.append(
        [None, None, None, None, None, None, None, None, None, 0, 0, 0, None, None, None, None]
    )
    wb.save(path)


@pytest.fixture
def mini_ventas(tmp_path) -> Path:
    path = tmp_path / "mini-ventas.xlsx"
    _mini_workbook(path)
    return path


# --------------------------------------------------------------------------- #
# Module-level DB cleanup (canonical tipos + test rows)
# --------------------------------------------------------------------------- #


def _ventas_id_de_test(db):
    """IDs de las Ventas creadas por este modulo (via detalle test o cliente)."""
    ids = set(
        db.query(DetalleVenta.venta_id)
        .filter(
            DetalleVenta.producto_id.in_(
                db.query(Producto.id).filter(
                    Producto.nombre.in_([P_SET, P_TOTE, f"{P} Venta Combo"])
                )
            )
        )
        .all()
    )
    cli = db.query(Cliente).filter(Cliente.nombre.in_([P_CLI, P_CLI2])).first()
    if cli is not None:
        ids |= set(db.query(Venta.id).filter(Venta.cliente_id == cli.id).all())
    return {i for (i,) in ids}


def _limpiar_ventas_test(db) -> None:
    """Borra las filas de venta relevantes al inicio de cada test (el fixture
    de modulo comparte sesion; aislamos por nombre exacto)."""
    for vid in _ventas_id_de_test(db):
        db.query(DetalleVenta).filter(DetalleVenta.venta_id == vid).delete(
            synchronize_session=False
        )
        db.query(Venta).filter(Venta.id == vid).delete(synchronize_session=False)
    db.query(Cliente).filter(Cliente.nombre.in_([P_CLI, P_CLI2])).delete(synchronize_session=False)
    db.commit()


def _borrar_filas_test(db) -> None:
    """Remove rows this test module injected (exact-name matches only)."""
    _limpiar_ventas_test(db)
    ids_productos_test = [
        i
        for (i,) in db.query(Producto.id)
        .filter(Producto.nombre.in_([P_SET, P_TOTE, f"{P} Venta Combo"]))
        .all()
    ]
    db.query(BomInsumo).filter(BomInsumo.producto_id.in_(ids_productos_test)).delete(
        synchronize_session=False
    )
    db.query(BomInsumo).filter(
        BomInsumo.insumo_id.in_(
            db.query(Insumo.id).filter(Insumo.nombre.in_([P_TELA, f"{P} Caja Empaque"]))
        )
    ).delete(synchronize_session=False)
    db.query(Insumo).filter(Insumo.nombre == P_TELA).delete(synchronize_session=False)
    db.query(Insumo).filter(Insumo.nombre == f"{P} Caja Empaque").delete(synchronize_session=False)
    db.query(VarianteProducto).filter(VarianteProducto.producto_id.in_(ids_productos_test)).delete(
        synchronize_session=False
    )
    db.query(Producto).filter(Producto.id.in_(ids_productos_test)).delete(synchronize_session=False)
    # Remove the canonical catalog tipos that bootstrap_catalogo() inserts.
    db.query(TipoProducto).filter(
        TipoProducto.nombre.in_(["Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"])
    ).delete(synchronize_session=False)
    db.commit()


@pytest.fixture(autouse=True, scope="module")
def _cleanup_after_module():
    yield
    db = SessionLocal()
    try:
        _borrar_filas_test(db)
    finally:
        db.close()


@pytest.fixture(scope="module")
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _preparar_catalogo(db) -> None:
    """F3-ish: bootstrap + producto/variante + insumo + BOM_Insumos (test)."""
    from migrate.catalog import bootstrap_catalogo, upsert_insumo, upsert_producto

    _limpiar_ventas_test(db)
    db.query(BomInsumo).filter(
        BomInsumo.insumo_id.in_(db.query(Insumo.id).filter(Insumo.nombre == P_TELA))
    ).delete(synchronize_session=False)
    db.query(Insumo).filter(Insumo.nombre == P_TELA).delete(synchronize_session=False)
    db.query(VarianteProducto).filter(
        VarianteProducto.producto_id.in_(
            db.query(Producto.id).filter(Producto.nombre.in_([P_SET, P_TOTE]))
        )
    ).delete(synchronize_session=False)
    db.query(Producto).filter(Producto.nombre.in_([P_SET, P_TOTE])).delete(
        synchronize_session=False
    )
    db.commit()

    bootstrap_catalogo(db)
    ids = {}
    ids[P_TELA] = upsert_insumo(db, P_TELA, categoria_nombre="Telas").id
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    tela.stock_actual = Decimal("10")
    ids[P_SET] = upsert_producto(db, P_SET, tipo="Accesorio", variantes=("S",)).id
    ids[P_TOTE] = upsert_producto(db, P_TOTE, tipo="Accesorio").id
    db.flush()
    set_prod = db.query(Producto).filter(Producto.nombre == P_SET).one()
    var = (
        db.query(VarianteProducto)
        .filter(
            VarianteProducto.producto_id == set_prod.id,
            VarianteProducto.nombre_variante == "S",
        )
        .one()
    )
    db.add(
        BomInsumo(
            producto_id=set_prod.id,
            insumo_id=ids[P_TELA],
            variante_id=var.id,
            cantidad_requerida=Decimal("2"),
            porcentaje_desperdicio=Decimal("0"),
        )
    )
    db.commit()


# --------------------------------------------------------------------------- #
# 1. Pure: plan from workbook (bounded, real layout, SCOPE OUT)
# --------------------------------------------------------------------------- #


def test_plan_ventas_lee_productos_reales(mini_ventas):
    from migrate.sales import plan_ventas

    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)

    assert plan.conteo_ventas == 2  # SCOPE OUT + junk excluidos
    set_ = [v for v in plan.ventas if v.producto_nombre == P_SET][0]
    assert set_.variante_nombre == "S"
    assert set_.precio == Decimal("71250.00")
    assert set_.costo == Decimal("26109")
    assert set_.fecha == datetime(2026, 3, 20, tzinfo=UTC)
    assert set_.cliente_nombre == P_CLI


def test_plan_ventas_marca_scope_out_y_junk(mini_ventas):
    from migrate.sales import plan_ventas

    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)

    assert plan.scope_out == 2  # R4 sin producto + R5 junk totales
    nombres = {v.producto_nombre for v in plan.ventas}
    assert "Olga" not in map(str, nombres)


def test_plan_ventas_fecha_precio_costo_no_inventados(mini_ventas):
    from migrate.sales import plan_ventas

    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)

    tote = [v for v in plan.ventas if v.producto_nombre == P_TOTE][0]
    assert tote.fecha == datetime(2026, 4, 24, tzinfo=UTC)
    assert tote.precio == Decimal("45000.00")
    assert tote.costo == Decimal("25765.09524")


# --------------------------------------------------------------------------- #
# 2. DB: INSERT directo Venta+Detalle con snapshot + canal 'feria'
# --------------------------------------------------------------------------- #


def test_aplicar_ventas_inserta_con_fecha_real_canal_y_snapshot(db, mini_ventas):
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)
    res = aplicar_ventas(db, plan, canal_venta="feria")
    db.commit()

    assert res["insertadas"] == 2
    ventas = db.query(Venta).all()
    assert len(ventas) == 2
    for v in ventas:
        assert v.canal_venta == "feria"  # decision producto
        assert v.estado == "completada"
        assert v.descuento_porcentaje == Decimal("0")  # sin doble descuento
        assert v.fecha is not None  # fecha real (nunca now())
    set_venta = (
        db.query(Venta)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .filter(Cliente.nombre == P_CLI)
        .one()
    )
    detalle = db.query(DetalleVenta).filter(DetalleVenta.venta_id == set_venta.id).one()
    # costo FULL del Excel (igual a H), sin recalcular con WAC actual
    assert detalle.costo_unitario_aplicado == Decimal("26109")
    assert detalle.precio_unitario_aplicado == Decimal("71250.00")
    assert set_venta.fecha.strftime("%Y-%m-%d") == "2026-03-20"


def test_aplicar_ventas_sin_doble_descuento(db, mini_ventas):
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)
    aplicar_ventas(db, plan, canal_venta="feria")
    db.commit()

    cli = db.query(Cliente).filter(Cliente.nombre == P_CLI).one()
    set_venta = db.query(Venta).filter(Venta.cliente_id == cli.id).one()
    # La nota DESC (DESC 25%) esta en la col O pero el precio ya viene
    # descontado del Excel (VTA-2): descuento_porcentaje=0, precio tal-cual.
    assert set_venta.descuento_porcentaje == Decimal("0")
    det = db.query(DetalleVenta).filter(DetalleVenta.venta_id == set_venta.id).one()
    assert det.precio_unitario_aplicado == Decimal("71250.00")


def test_aplicar_ventas_destock_batch(db, mini_ventas):
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    assert tela.stock_actual == Decimal("10")
    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)
    aplicar_ventas(db, plan, canal_venta="feria")  # boom + destock in lote
    db.commit()

    # Set (1 und) con BOM requerida 2 m -> explosion 2 (desperdicio 0);
    # Tote sin BOM -> sin consumo.
    db.refresh(tela)
    assert tela.stock_actual == Decimal("8")


def test_aplicar_ventas_idempotente(db, mini_ventas):
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)
    aplicar_ventas(db, plan, canal_venta="feria")
    db.commit()
    aplicar_ventas(db, plan, canal_venta="feria")  # re-ejecucion
    db.commit()

    assert db.query(Venta).count() == 2  # no duplica
    assert db.query(DetalleVenta).count() == 2


def test_aplicar_ventas_cliente_upsert(db, mini_ventas):
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)
    aplicar_ventas(db, plan, canal_venta="feria")
    db.commit()

    cli = db.query(Cliente).filter(Cliente.nombre == P_CLI).one()
    assert db.query(Venta).filter(Venta.cliente_id == cli.id).count() == 1


# --------------------------------------------------------------------------- #
# 3. Rollback: stock insuficiente -> 409, cero residuos
# --------------------------------------------------------------------------- #


def test_aplicar_ventas_stock_insuficiente_rollback(db, mini_ventas):
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    tela.stock_actual = Decimal("1")  # alcanza solo 0.5 und del BOM (2 m)
    db.commit()

    with LibroMigracion(mini_ventas) as libro:
        plan = plan_ventas(libro)
    with pytest.raises(InsufficientStockError) as exc:
        aplicar_ventas(db, plan, canal_venta="feria")
    assert exc.value.status_code == 409
    db.rollback()
    # EXM-4: ninguna venta de la fase queda persistida (rollback de fase)
    assert db.query(Venta).count() == 0
    assert db.query(DetalleVenta).count() == 0


def test_aplicar_ventas_no_descuenta_empaques_de_combo(db, tmp_path):
    """Regression (2026-08): los empaques de los combos (Caja/Vela/Papel/Envio,
    categoria Empaques) son consumibles SIN inventario rastreable: no estan en
    OCT25 ni tienen compras WAC. F5 debe excluirlos del destock o la venta de
    un combo (CAJA SACA LAS GARRAS) falla con InsufficientStockError aun con
    stock real suficiente de los materiales."""
    from migrate.catalog import upsert_insumo, upsert_producto
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    # Combo con BOM: tela (stock 10) + empaque Caja (stock 0, categoria Empaques)
    caja = upsert_insumo(db, f"{P} Caja Empaque", categoria_nombre="Empaques")
    combo = upsert_producto(db, f"{P} Venta Combo", tipo="Combo")
    db.flush()
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    db.add(
        BomInsumo(
            producto_id=combo.id,
            insumo_id=tela.id,
            cantidad_requerida=Decimal("2"),
            porcentaje_desperdicio=Decimal("0"),
        )
    )
    db.add(
        BomInsumo(
            producto_id=combo.id,
            insumo_id=caja.id,
            cantidad_requerida=Decimal("1"),
            porcentaje_desperdicio=Decimal("0"),
        )
    )
    db.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VENTAS"
    ws.append(
        [
            "Producto",
            None,
            None,
            None,
            None,
            None,
            "Precio Venta",
            "Costo",
            "Ganancias",
            None,
            None,
            None,
            "Fecha",
            None,
            None,
            "Cliente",
        ]
    )
    ws.append(
        [
            f"{P} Venta Combo",
            None,
            None,
            None,
            None,
            None,
            295000.0,
            129388.0,
            165612.0,
            None,
            None,
            None,
            datetime(2026, 7, 13),
            None,
            None,
            "gaby",
        ]
    )
    path = tmp_path / "mini-combo.xlsx"
    wb.save(path)

    with LibroMigracion(path) as libro:
        plan = plan_ventas(libro)
    aplicar_ventas(db, plan, canal_venta="feria")  # no debe lanzar
    db.commit()

    db.refresh(tela)
    db.refresh(caja)
    assert tela.stock_actual == Decimal("8")  # tela SI se descuenta
    assert caja.stock_actual == Decimal("0")  # empaque NO se descuenta


def test_aplicar_ventas_permitir_deficit_alerta_y_no_rollback(db, tmp_path):
    """Decision de negocio (2026-08): el historico de ventas supera el
    inventario comprado (Lino vertigo 4.8m vs 1.8m, Satin elastico 4.8 vs 2,
    Tela a cuadros 1.8 vs 1) -- el negocio vendio de stock previo no
    registrado. Con permitir_deficit=True F5 registra las ventas, descuenta
    dejando stock NEGATIVO y reporta WARN por insumo deficitario, sin
    rollback (EXM-4 relajado para la migracion historica)."""
    from migrate.sales import aplicar_ventas, plan_ventas

    _preparar_catalogo(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    tela.stock_actual = Decimal("1")  # BOM requiere 2 m -> deficit
    db.commit()

    path = tmp_path / "mini-ventas-deficit.xlsx"
    _mini_workbook(path)
    with LibroMigracion(path) as libro:
        plan = plan_ventas(libro)
    res = aplicar_ventas(db, plan, canal_venta="feria", permitir_deficit=True)
    db.commit()

    db.refresh(tela)
    assert tela.stock_actual == Decimal("-1")  # deficit permitido (stock negativo)
    assert db.query(Venta).count() == 2  # las ventas SI se persistieron
    assert db.query(DetalleVenta).count() == 2
    assert res["destock"] >= 1


# --------------------------------------------------------------------------- #
# 4. Phase: runner registry + real workbook dry run (NFR-2)
# --------------------------------------------------------------------------- #


def test_f5_registrada_en_runner():
    from migrate import FASE_RUNNERS, FASES, FASES_IMPLEMENTADAS

    assert any(f.id == "F5" for f in FASES)
    assert "F5" in FASE_RUNNERS
    assert "F5" in FASES_IMPLEMENTADAS


# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# 5. Variantes XXS-XL: omit talla-less sized rows (MIG-3) + NULL-matching (MIG-4)
# --------------------------------------------------------------------------- #


def test_aplicar_ventas_omitida_sin_talla_no_estalla(tmp_path):
    """MIG-3/D2: una venta de un producto TALLADO sin talla en la fila NO se
    inserta (res['omitidas'] == 1), la otra fila si, y la fase NO estalla con
    DomainValidationError (el omit ocurre ANTES de la explosion de inventory)."""
    from migrate.sales import aplicar_ventas, plan_ventas

    db = SessionLocal()
    try:
        _preparar_catalogo(db)  # P_SET con variantes ("S",), P_TOTE sin variantes

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VENTAS"
        ws.append(
            [
                "Producto",
                None,
                None,
                None,
                None,
                None,
                "Precio Venta",
                "Costo",
                "Ganancias",
                None,
                None,
                None,
                "Fecha",
                None,
                None,
                "Cliente",
            ]
        )
        # P_SET (producto tallado) SIN talla -> omitida MIG-3.
        ws.append(
            [
                P_SET,
                None,
                None,
                None,
                None,
                None,
                71250.0,
                26109,
                45141,
                None,
                None,
                None,
                datetime(2026, 3, 20),
                None,
                "DESC 25%",
                P_CLI,
            ]
        )
        # P_TOTE (sin variantes) sin talla -> se inserta igual.
        ws.append(
            [
                P_TOTE,
                None,
                None,
                None,
                None,
                None,
                45000.0,
                25765.09524,
                19234.90476,
                None,
                None,
                None,
                datetime(2026, 4, 24),
                None,
                None,
                P_CLI2,
            ]
        )
        path = tmp_path / "mini-sin-talla.xlsx"
        wb.save(path)

        with LibroMigracion(path) as libro:
            plan = plan_ventas(libro)
        res = aplicar_ventas(db, plan, canal_venta="feria")  # no debe lanzar
        db.commit()

        assert res["omitidas"] == 1
        assert res["insertadas"] == 1  # solo P_TOTE
        set_prod = db.query(Producto).filter(Producto.nombre == P_SET).one()
        assert (
            db.query(DetalleVenta).filter(DetalleVenta.producto_id == set_prod.id).count() == 0
        )
        assert db.query(Venta).count() == 1
    finally:
        db.rollback()
        db.close()


def test_aplicar_ventas_rerun_matchea_fila_null_historica(tmp_path):
    """MIG-4/D3: idempotencia unidireccional — una linea del plan CON talla
    matchea la fila NULL historica (insertada por el codigo viejo sin
    variantes): re-run -> ya_presentes, 0 duplicados."""
    from migrate.sales import aplicar_ventas, plan_ventas

    db = SessionLocal()
    try:
        _preparar_catalogo(db)
        set_prod = db.query(Producto).filter(Producto.nombre == P_SET).one()
        cli = Cliente(nombre=P_CLI)
        db.add(cli)
        db.flush()
        venta_historica = Venta(
            fecha=datetime(2026, 3, 20, tzinfo=UTC),
            cliente_id=cli.id,
            canal_venta="feria",
            descuento_porcentaje=Decimal("0"),
            estado="completada",
            total_venta=Decimal("71250.00"),
        )
        db.add(venta_historica)
        db.flush()
        db.add(
            DetalleVenta(
                venta_id=venta_historica.id,
                producto_id=set_prod.id,
                variante_id=None,  # fila historica sin variante
                cantidad=Decimal("1"),
                precio_unitario_aplicado=Decimal("71250.00"),
                costo_unitario_aplicado=Decimal("26109"),
            )
        )
        db.commit()

        path = tmp_path / "mini-null.xlsx"
        _mini_workbook(path)  # P_SET CON talla "S" + P_TOTE
        with LibroMigracion(path) as libro:
            plan = plan_ventas(libro)
        res = aplicar_ventas(db, plan, canal_venta="feria")
        db.commit()

        assert res["insertadas"] == 1  # solo P_TOTE
        assert res["ya_presentes"] == 1  # la linea "S" matcheo la NULL historica
        assert db.query(Venta).count() == 2  # sin duplicados
        assert db.query(DetalleVenta).count() == 2
    finally:
        db.rollback()
        db.close()


def test_aplicar_ventas_combo_sin_variante_no_matchea_fila_con_variante(tmp_path):
    """MIG-4/D3 (guard unidireccional): un combo sin talla (plan None) NO
    matchea una fila DB que SI tiene variante -> el combo se inserta; y el
    omit MIG-3 NO aplica a combos (tipo == 'Combo')."""
    from migrate.catalog import upsert_producto
    from migrate.sales import aplicar_ventas, plan_ventas

    db = SessionLocal()
    try:
        _preparar_catalogo(db)
        set_prod = db.query(Producto).filter(Producto.nombre == P_SET).one()
        set_var = (
            db.query(VarianteProducto)
            .filter(
                VarianteProducto.producto_id == set_prod.id,
                VarianteProducto.nombre_variante == "S",
            )
            .one()
        )
        combo = upsert_producto(db, f"{P} Venta Combo", tipo="Combo")
        db.flush()
        cli = Cliente(nombre=P_CLI)
        db.add(cli)
        db.flush()
        venta_rara = Venta(
            fecha=datetime(2026, 7, 13, tzinfo=UTC),
            cliente_id=cli.id,
            canal_venta="feria",
            descuento_porcentaje=Decimal("0"),
            estado="completada",
            total_venta=Decimal("295000"),
        )
        db.add(venta_rara)
        db.flush()
        db.add(
            DetalleVenta(
                venta_id=venta_rara.id,
                producto_id=combo.id,
                variante_id=set_var.id,  # dato historico raro: fila de combo con variante
                cantidad=Decimal("1"),
                precio_unitario_aplicado=Decimal("295000"),
                costo_unitario_aplicado=Decimal("129388"),
            )
        )
        db.commit()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VENTAS"
        ws.append(
            [
                "Producto",
                None,
                None,
                None,
                None,
                None,
                "Precio Venta",
                "Costo",
                "Ganancias",
                None,
                None,
                None,
                "Fecha",
                None,
                None,
                "Cliente",
            ]
        )
        ws.append(
            [
                f"{P} Venta Combo",
                None,
                None,
                None,
                None,
                None,
                295000.0,
                129388.0,
                165612.0,
                None,
                None,
                None,
                datetime(2026, 7, 13),
                None,
                None,
                P_CLI,
            ]
        )
        path = tmp_path / "mini-combo-variante.xlsx"
        wb.save(path)

        with LibroMigracion(path) as libro:
            plan = plan_ventas(libro)
        res = aplicar_ventas(db, plan, canal_venta="feria")
        db.commit()

        assert res["insertadas"] == 1  # el combo se re-inserta (None != variante)
        assert res["ya_presentes"] == 0
        assert res["omitidas"] == 0  # combos nunca se omiten por MIG-3
    finally:
        db.rollback()
        db.close()


def test_cargar_ventas_dry_run_real_no_escribe():
    from migrate.sales import cargar_ventas

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    db = SessionLocal()
    try:
        antes_ventas = db.query(Venta).count()
        antes_det = db.query(DetalleVenta).count()
        ctx = MigrationContext.para_fase(FaseOptions(source=REAL_XLSX, modo="dry-run"), "F5")
        plan = cargar_ventas(ctx)
        # NFR-2: dry-run write nothing.
        assert db.query(Venta).count() == antes_ventas
        assert db.query(DetalleVenta).count() == antes_det
        # The recalculated workbook's VENTAS sheet has #VALUE! on J/K/L and is
        # NOT usable: F5 reads csv/ARPIA - VENTAS.csv as its source. The CSV
        # has header at line 0 and 21 real data rows (lines 1..21), all with
        # product -> 21 historical sales (was 13 from the legacy sheet).
        assert plan.conteo_ventas == 21
        assert not ctx.report.tenga_errores
    finally:
        db.close()


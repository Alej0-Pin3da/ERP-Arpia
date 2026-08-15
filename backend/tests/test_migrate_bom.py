"""Contract tests for migrate.bom - F3 BOM phase (PR#5 slice).

Covers STRICT TDD acceptance from tasks #424 T6 (spec BOM-1..BOM-3,
EXM-2/3/4, NFR-1/2; design #423 bom.py + D4):

- Pure: BOM quantity conversion from the Excel ''cantidad Cms'' cell (LINEAR
  cm of consumption per garment) to the insumo canonical unit: Telas -> m via
  metros = cm / 100 (the material width does NOT participate), Herrajes cm2 ->
  cm2 as-is, 'un' piece counts as-is; junk rows filtered.
- Workbook -> plan: per BOM sheet the LEFT block feeds the product and, where
  the sheet defines TWO products (BLUSAS: MANGA LARGA + MANGA CORTA), the
  right block feeds the second product; TANGA ghost sub-blocks (sheets whose
  right block is empty) are skipped.
- DB (real PostgreSQL): applying creates BomInsumo rows with the exact
  converted quantity and variante_id NULL, and is idempotent on re-run (PG
  UNIQUE does not apply over NULLs -> manual dedup by natural key); combos
  create BomProducto rows by (combo_id, producto_incluido_id) with
  skip-si-existe.
- Phase: F3 registered in the runner registry; dry-run on the real workbook
  writes 0 rows (NFR-2).

Test-injected rows use the 'Migratest ' prefix so cleanup never touches real
migration data; the canonical catalog tipos inserted by bootstrap_catalogo()
are removed at module cleanup (same pattern as test_migrate_catalog /
test_migrate_purchases).
"""

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.db.session import SessionLocal
from app.models import BomInsumo, BomProducto, Insumo, Producto, TipoProducto
from migrate.context import FaseOptions, MigrationContext
from migrate.loaders import LibroMigracion

REAL_XLSX = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")

# Names injected by tests; cleanup deletes exactly these, never catalog rows.
P = "Migratest"
P_TELA = f"{P} BOM Tela"
P_ARG = f"{P} BOM Argolla"
P_TUL = f"{P} BOM Tul"
P_CORSET = f"{P} BOM Corset"
P_BLUSA_ML = f"{P} BOM Blusa ML"
P_BLUSA_MC = f"{P} BOM Blusa MC"
P_COMBO = f"{P} BOM Combo"
P_CADENA = f"{P} BOM Cadena"

# Canonical catalog insumos (recalculated 2026-08 workbook) that recipe names
# resolve to via ALIASES_BOM_A_CATALOGO (now mostly exact match / identity).
# The orchestrator creates these in the real DB; tests create them in the fixture.
P_ALIAS_ARG = "Argolla 10 mm"
P_ALIAS_TIRA = "Tira de brasier"
P_ALIAS_POWERNET = "Powernet negro delgado (corsets)"
# Recipe material with neither alias nor catalog match -> omitted.
P_FANTASMA = f"{P} BOM Material Fantasma"


# --------------------------------------------------------------------------- #
# Mini workbook builders
# --------------------------------------------------------------------------- #


def _mini_workbook(path: Path) -> None:
    """Mini workbook: CORSET (left block + ghost TANGA right), BLUSAS (two real
    products), CAJAS (one combo). Rows within SHEET_BOUNDS."""
    wb = openpyxl.Workbook()

    corset = wb.active
    corset.title = "CORSET"
    corset.append(["CORSET", None, None, None, None, None, None, None, "TANGA"])
    corset.append(["Producto", "cantidad Cms", "Ancho", "area", "valor metro", "valor total"])
    # Layout REAL alineado (2026-08): B=cantidad lineal, C=ancho, D=area=B*C.
    # R3..R5: left block (CORSET product). R9: right TANGA ghost (skipped).
    corset.append([P_TELA, 64, 37, 2368, 2.576888889, None])
    corset.append([P_ARG, 2, 1, 2, 72, None])
    corset.append([P_TUL, 24, 18, 432, 4.5, None])
    corset.append([4.0, None, None, None, None, None])  # junk numeric
    corset.append(["Horas trabajo", None, None, None, None, None])  # junk
    corset.cell(row=9, column=9, value=P_TELA)  # I9 ghost TANGA
    corset.cell(row=9, column=10, value=200)
    corset.cell(row=9, column=11, value=1)
    corset.cell(row=9, column=12, value=200)

    blusas = wb.create_sheet("BLUSAS")
    blusas.append(["MANGA LARGA", None] * 4 + ["MANGA CORTA"])
    blusas.append(["Producto", "cantidad Cms", "Ancho", "area", "valor metro", "valor total"])
    # Layout REAL: izquierda A/B (nombre/cantidad), derecha I/J (nombre/cantidad).
    blusas.append([P_TELA, 45, 54, 2430, 2.5, None])  # R3 left (ML): 45 cm
    blusas.append([P_ARG, 4, 1, 4, 72, None])  # R4 left (ML): 4 piezas
    blusas.cell(row=3, column=9, value=P_TELA)  # I3 right (MC): 60 cm
    blusas.cell(row=3, column=10, value=60)
    blusas.cell(row=3, column=11, value=120)
    blusas.cell(row=3, column=12, value=7200)

    cajas = wb.create_sheet("CAJAS")
    # Real layout: combo name at R2 (B2), headers R3, data R4+ (SHEET_BOUNDS 4..13)
    cajas.cell(row=2, column=2, value=P_COMBO)  # B2 combo name
    cajas.append([None, "PRENDAS", "COSTO", "PRECIO", None, None])
    cajas.append([None, P_CORSET, 38804.62508, 110000])
    cajas.append([None, P_BLUSA_ML, 21561.75289, 90000])
    wb.save(path)


@pytest.fixture
def mini_bom(tmp_path) -> Path:
    path = tmp_path / "mini-bom.xlsx"
    _mini_workbook(path)
    return path


def _mini_workbook_aliases(path: Path) -> None:
    """Mini workbook: CORSET with recipe names that resolve via aliases plus one
    phantom material (no alias, no catalog match) that must be omitted."""
    wb = openpyxl.Workbook()
    corset = wb.active
    corset.title = "CORSET"
    corset.append(["CORSET", None, None, None, None, None, None, None, "TANGA"])
    corset.append(["Producto", "Ancho", "Alto", "cantidad Cms", "valor metro", "valor total"])
    corset.append(["Argolla 10 mm", 2, 1, 2, 72, None])
    corset.append(["Tira de brasier", 1, 1, 1, 3000, None])
    corset.append(["Powernet negro delgado (corsets)", 1, 1, 1, 12000, None])
    corset.append([P_FANTASMA, 1, 1, 1, 999, None])
    wb.save(path)


@pytest.fixture
def mini_bom_aliases(tmp_path) -> Path:
    path = tmp_path / "mini-bom-aliases.xlsx"
    _mini_workbook_aliases(path)
    return path


# --------------------------------------------------------------------------- #
# Module-level DB cleanup (canonical tipos + test rows)
# --------------------------------------------------------------------------- #


def _borrar_filas_test(db) -> None:
    """Remove rows this test module injected (exact-name matches only)."""
    insumo_nombres = [P_TELA, P_ARG, P_TUL, P_ALIAS_ARG, P_ALIAS_TIRA, P_ALIAS_POWERNET]
    insumos = db.query(Insumo).filter(Insumo.nombre.in_(insumo_nombres)).all()
    for insumo in insumos:
        db.query(BomInsumo).filter(BomInsumo.insumo_id == insumo.id).delete(
            synchronize_session=False
        )
    db.query(Insumo).filter(Insumo.nombre.in_(insumo_nombres)).delete(synchronize_session=False)
    productos = (
        db.query(Producto)
        .filter(Producto.nombre.in_([P_CORSET, P_BLUSA_ML, P_BLUSA_MC, P_COMBO]))
        .all()
    )
    for prod in productos:
        db.query(BomInsumo).filter(BomInsumo.producto_id == prod.id).delete(
            synchronize_session=False
        )
        db.query(BomProducto).filter(BomProducto.combo_id == prod.id).delete(
            synchronize_session=False
        )
        db.query(BomProducto).filter(BomProducto.producto_incluido_id == prod.id).delete(
            synchronize_session=False
        )
    db.query(Producto).filter(
        Producto.nombre.in_([P_CORSET, P_BLUSA_ML, P_BLUSA_MC, P_COMBO])
    ).delete(synchronize_session=False)
    # Remove the canonical catalog tipos that bootstrap_catalogo() inserts.
    # They are migration content, not app seed data; leaving them pollutes the
    # shared per-sheet DB and breaks pagination tests that assume an empty table.
    db.query(TipoProducto).filter(
        TipoProducto.nombre.in_(["Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"])
    ).delete(synchronize_session=False)
    db.commit()


@pytest.fixture(autouse=True, scope="module")
def _cleanup_after_module():
    yield
    db = SessionLocal()
    try:
        _borrar_filas_test(db)
    finally:
        db.close()


@pytest.fixture(scope="module")
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _preparar_catalogo(db) -> dict[str, int]:
    """F1-ish: bootstrap + insumos BOM del mini + productos de prueba."""
    from migrate.catalog import bootstrap_catalogo, upsert_insumo, upsert_producto

    bootstrap_catalogo(db)
    db.flush()
    ids = {}
    ids[P_TELA] = upsert_insumo(db, P_TELA, categoria_nombre="Telas").id
    ids[P_ARG] = upsert_insumo(db, P_ARG, categoria_nombre="Herrajes").id
    ids[P_TUL] = upsert_insumo(db, P_TUL, categoria_nombre="Telas").id
    ids[P_CORSET] = upsert_producto(db, P_CORSET, tipo="Lencería").id
    ids[P_BLUSA_ML] = upsert_producto(db, P_BLUSA_ML, tipo="Blusa").id
    ids[P_BLUSA_MC] = upsert_producto(db, P_BLUSA_MC, tipo="Blusa").id
    ids[P_COMBO] = upsert_producto(db, P_COMBO, tipo="Combo").id
    db.commit()
    return ids


def _preparar_catalogo_aliases(db) -> dict[str, int]:
    """Like _preparar_catalogo plus the canonical catalog insumos that the
    recipe alias names (Argolla 10 mm, ...) resolve to in F3 apply."""
    ids = _preparar_catalogo(db)
    from migrate.catalog import upsert_insumo

    ids[P_ALIAS_ARG] = upsert_insumo(db, P_ALIAS_ARG, categoria_nombre="Herrajes").id
    ids[P_ALIAS_TIRA] = upsert_insumo(db, P_ALIAS_TIRA, categoria_nombre="Telas").id
    ids[P_ALIAS_POWERNET] = upsert_insumo(db, P_ALIAS_POWERNET, categoria_nombre="Telas").id
    db.commit()
    return ids


# --------------------------------------------------------------------------- #
# 1. Pure: quantity conversion (cm lineales -> canonical unit)
# --------------------------------------------------------------------------- #


def test_convertir_cantidad_bom_m_centimetros_lineales():
    from migrate.bom import convertir_cantidad_bom

    # 'cantidad Cms' = centimetros LINEALES de consumo por prenda.
    # metros = cm / 100; el ancho del material no participa.
    assert convertir_cantidad_bom("5", "Elastico pitillo rosa", "m") == Decimal("0.05")
    assert convertir_cantidad_bom("74", "Tira de Brasier negro 10 mts", "m") == Decimal("0.74")
    assert convertir_cantidad_bom("6", "Sesgo Elastico 10 mts", "m") == Decimal("0.06")
    assert convertir_cantidad_bom("40", "Framilon elastico plano 20 mts", "m") == Decimal("0.40")


def test_convertir_cantidad_bom_m_ignora_ancho_del_nombre():
    from migrate.bom import convertir_cantidad_bom

    # El ancho declarado en el nombre ('19 cm') YA NO afecta la conversion:
    # el consumo es lineal, metros = cm / 100 (regresion del bug 100x).
    assert convertir_cantidad_bom("74", "Encaje Elastico 19 cm negro 10 mts", "m") == Decimal(
        "0.74"
    )


def test_convertir_cantidad_bom_piezas_y_cm2():
    from migrate.bom import convertir_cantidad_bom

    # Herrajes en "un": la cantidad Cms es conteo de piezas (2 argollas)
    assert convertir_cantidad_bom("2", "Argolla 10 mm", "un") == Decimal("2")
    # Herrajes en "cm2" (precio por cm2): cantidad se conserva en cm2
    assert convertir_cantidad_bom("4670", "Sublimacion (cm2)", "cm2") == Decimal("4670")


def test_convertir_cantidad_bom_rechaza_no_numerico():
    from migrate.bom import convertir_cantidad_bom

    assert convertir_cantidad_bom(None, "Tela", "m") is None
    assert convertir_cantidad_bom("#DIV/0!", "Tela", "m") is None
    assert convertir_cantidad_bom("0", "Tela", "m") is None  # cero no es consumo
    assert convertir_cantidad_bom("-3", "Tela", "m") is None  # negativo tampoco


# --------------------------------------------------------------------------- #
# 2. Workbook -> plan (bloques izquierdo/derecho, ghost TANGA, combos)
# --------------------------------------------------------------------------- #


def _bloques_mini():
    """Mini-book product mapping (hoja -> producto catálogo de prueba)."""
    from migrate.bom import BLOQUES_BOM

    bloques = dict(BLOQUES_BOM)
    bloques["CORSET"] = (P_CORSET, None)  # right block TANGA = ghost
    bloques["BLUSAS"] = (P_BLUSA_ML, P_BLUSA_MC)
    return bloques


def test_plan_bom_asigna_producto_por_hoja(mini_bom):
    from migrate.bom import plan_bom

    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())

    productos = {item.producto_nombre for item in plan.insumos}
    # La hoja CORSET alimenta el producto Corset; BLUSAS alimenta ambos productos
    assert P_CORSET in productos
    assert P_BLUSA_ML in productos
    assert P_BLUSA_MC in productos


def test_plan_bom_descarta_ghost_tanga_y_junk(mini_bom):
    from migrate.bom import plan_bom

    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())

    # La fila fantasma TANGA (I9) NO alimenta el Corset (2xTela? no: solo izq)
    corset = [item for item in plan.insumos if item.producto_nombre == P_CORSET]
    assert len(corset) == 3  # Tela + Argolla + Tul; junk y TANGA excluidos
    assert all(item.cantidad > 0 for item in corset)


def test_plan_bom_cantidades_convertidas_exactas(mini_bom):
    from migrate.bom import plan_bom

    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())

    por = {(item.producto_nombre, item.insumo_nombre): item.cantidad for item in plan.insumos}
    # Layout real (2026-08): cantidad LINEAL en col B -> Corset Tul 24 cm / 100 = 0.24 m
    assert por[(P_CORSET, P_TUL)] == Decimal("0.24")
    # Corset Tela: 64 cm lineales / 100 = 0.64 m (el ancho ya no participa)
    assert por[(P_CORSET, P_TELA)] == Decimal("0.64")
    # Corset Argolla (un): 2 piezas
    assert por[(P_CORSET, P_ARG)] == Decimal("2")
    # Blusa ML (bloque izquierdo): 45 / 100 = 0.45 m
    assert por[(P_BLUSA_ML, P_TELA)] == Decimal("0.45")
    # Blusa MC (bloque derecho, col J): 60 / 100 = 0.6 m
    assert por[(P_BLUSA_MC, P_TELA)] == Decimal("0.6")


def test_plan_bom_combos_cajas(mini_bom):
    from migrate.bom import plan_bom

    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())

    combos = [c for c in plan.combos if c.combo_nombre == P_COMBO]
    assert len(combos) == 2  # Corset + Blusa ML del mini CAJAS
    assert {c.producto_incluido for c in combos} == {P_CORSET, P_BLUSA_ML}


def _mini_workbook_lote(path: Path) -> None:
    """Mini TOTEBAG con lote de 6 (layout real: 'UNIDAD | TOTAL 6 totebag').
    El consumo de la hoja es para el LOTE COMPLETO, no por unidad.
    SHEET_BOUNDS lee TOTEBAG desde R3, asi que los datos van en R3+.
    Layout real: B = cantidad lineal (cm)."""
    wb = openpyxl.Workbook()
    tot = wb.active
    tot.title = "TOTEBAG"
    tot.cell(row=3, column=1, value=P_TELA)
    tot.cell(row=3, column=2, value=53)  # 53 cm tela para 6 totebags
    tot.cell(row=4, column=1, value=P_ARG)
    tot.cell(row=4, column=2, value=12)  # 12 argollas para 6 totebags
    tot.cell(row=5, column=1, value="UNIDAD")
    tot.cell(row=5, column=2, value="TOTAL 6 totebag")
    tot.cell(row=6, column=1, value="COSTO TOTAL CONJUNTO")
    wb.save(path)


def test_plan_bom_lote_divido_cantidad_por_unidad(tmp_path):
    """Regression (2026-08, workbook real): las hojas BOM expresan consumo por
    LOTE ('UNIDAD | TOTAL 6 totebag', 'TOTAL 12/15 Prendas', 'TOTAL 8
    Prendas'), no por unidad. Antes del fix el BOM de Tote Bag quedaba 6x
    inflado (Cadena gris 53 cm -> 0.53 m por totebag), el destock F5 pedia
    6x lo real y fallaba con InsufficientStockError. Con lote 6: 53 cm / 6
    = 8.833 cm = 0.08833 m por unidad."""
    from migrate.bom import plan_bom

    path = tmp_path / "mini-lote.xlsx"
    _mini_workbook_lote(path)
    with LibroMigracion(path) as libro:
        plan = plan_bom(
            libro,
            bloques={"TOTEBAG": ("Tote Bag Arpia", None)},
            _lotes={"TOTEBAG": (6, None)},
        )

    por = {(i.producto_nombre, i.insumo_nombre): i.cantidad for i in plan.insumos}
    # Tela: 53 cm / 100 = 0.53 m, dividido por lote 6 -> 0.088333 m por unidad
    assert por[("Tote Bag Arpia", P_TELA)] == Decimal("0.08833333333333333333333333333")
    # Argolla (un): 12 piezas para 6 totebags -> 2 por unidad
    assert por[("Tote Bag Arpia", P_ARG)] == Decimal("2")


# --------------------------------------------------------------------------- #
# 3. DB: aplicar BOM + idempotencia (variante NULL) + multinivel
# --------------------------------------------------------------------------- #


def test_aplicar_bom_crea_insumos_con_cantidades(db, mini_bom):
    from migrate.bom import aplicar_bom, plan_bom

    _preparar_catalogo(db)
    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())
    aplicar_bom(db, plan)
    db.commit()

    corset = db.query(Producto).filter(Producto.nombre == P_CORSET).one()
    lineas = db.query(BomInsumo).filter(BomInsumo.producto_id == corset.id).all()
    assert len(lineas) == 3
    por = {item.insumo.nombre: item for item in lineas}
    # Layout real (2026-08): cantidad LINEAL en col B -> Tul 24 cm = 0.24 m
    assert por[P_TUL].cantidad_requerida == Decimal("0.24")
    assert por[P_TELA].cantidad_requerida == Decimal("0.64")
    assert all(item.variante_id is None for item in lineas)  # variante NULL


def test_aplicar_bom_idempotente_variante_null(db, mini_bom):
    from migrate.bom import aplicar_bom, plan_bom

    _preparar_catalogo(db)
    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())
    aplicar_bom(db, plan)
    db.commit()
    aplicar_bom(db, plan)
    db.commit()

    corset = db.query(Producto).filter(Producto.nombre == P_CORSET).one()
    lineas = db.query(BomInsumo).filter(BomInsumo.producto_id == corset.id).all()
    assert len(lineas) == 3  # re-ejecucion no duplica (dedup manual NULL)


def test_aplicar_bom_multinivel_combo(db, mini_bom):
    from migrate.bom import aplicar_bom, plan_bom

    _preparar_catalogo(db)
    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())
    aplicar_bom(db, plan)
    db.commit()

    combo = db.query(Producto).filter(Producto.nombre == P_COMBO).one()
    items = db.query(BomProducto).filter(BomProducto.combo_id == combo.id).all()
    assert len(items) == 2
    incluidos = {it.producto_incluido.nombre for it in items}
    assert incluidos == {P_CORSET, P_BLUSA_ML}
    # Los productos del combo tienen su propio BOM_Insumos (multinivel real)
    for it in items:
        assert (
            db.query(BomInsumo).filter(BomInsumo.producto_id == it.producto_incluido_id).count()
            >= 1
        )


def test_aplicar_bom_combo_idempotente(db, mini_bom):
    from migrate.bom import aplicar_bom, plan_bom

    _preparar_catalogo(db)
    with LibroMigracion(mini_bom) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())
    aplicar_bom(db, plan)
    db.commit()
    aplicar_bom(db, plan)
    db.commit()

    combo = db.query(Producto).filter(Producto.nombre == P_COMBO).one()
    assert db.query(BomProducto).filter(BomProducto.combo_id == combo.id).count() == 2


def test_aplicar_bom_alias_resuelve_insumos_canonicos(db, mini_bom_aliases):
    """Nombres cortos de receta -> insumo canonico del catalogo (ALIASES_BOM)."""
    from migrate.bom import aplicar_bom, plan_bom

    ids = _preparar_catalogo_aliases(db)
    with LibroMigracion(mini_bom_aliases) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())
    aplicar_bom(db, plan)
    db.commit()

    corset = db.query(Producto).filter(Producto.nombre == P_CORSET).one()
    por = {
        item.insumo.nombre: item
        for item in db.query(BomInsumo).filter(BomInsumo.producto_id == corset.id).all()
    }
    # 'Argolla 10 mm' recipe -> 'Argolla 10 mm' catalog insumo (now identity)
    assert P_ALIAS_ARG in por and por[P_ALIAS_ARG].insumo_id == ids[P_ALIAS_ARG]
    # 'Tira de brasier' recipe -> 'Tira de brasier' catalog insumo (now identity)
    assert P_ALIAS_TIRA in por and por[P_ALIAS_TIRA].insumo_id == ids[P_ALIAS_TIRA]
    # 'Powernet negro delgado (corsets)' recipe -> same-name catalog insumo (now identity)
    assert P_ALIAS_POWERNET in por and por[P_ALIAS_POWERNET].insumo_id == ids[P_ALIAS_POWERNET]


def test_aplicar_bom_sin_alias_ni_match_exacto_omite(db, mini_bom_aliases):
    """Material sin alias y sin match exacto sigue omitiendose (comport. actual)."""
    from migrate.bom import aplicar_bom, plan_bom

    _preparar_catalogo_aliases(db)
    with LibroMigracion(mini_bom_aliases) as libro:
        plan = plan_bom(libro, bloques=_bloques_mini())
    res = aplicar_bom(db, plan)
    db.commit()

    assert res["omitidos"] == 1  # solo el material fantasma
    corset = db.query(Producto).filter(Producto.nombre == P_CORSET).one()
    nombres = {
        item.insumo.nombre
        for item in db.query(BomInsumo).filter(BomInsumo.producto_id == corset.id).all()
    }
    assert P_FANTASMA not in nombres


# --------------------------------------------------------------------------- #
# 4. Phase: runner registry + real workbook dry run (NFR-2)
# --------------------------------------------------------------------------- #


def test_f3_registrada_en_runner():
    from migrate import FASE_RUNNERS, FASES, FASES_IMPLEMENTADAS

    assert any(f.id == "F3" for f in FASES)
    assert "F3" in FASE_RUNNERS
    assert "F3" in FASES_IMPLEMENTADAS


def test_cargar_bom_dry_run_real_no_escribe():
    from migrate.bom import cargar_bom

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    db = SessionLocal()
    try:
        antes_bom = db.query(BomInsumo).count()
        antes_combo = db.query(BomProducto).count()
        ctx = MigrationContext.para_fase(FaseOptions(source=REAL_XLSX, modo="dry-run"), "F3")
        cargar_bom(ctx)
        assert db.query(BomInsumo).count() == antes_bom
        assert db.query(BomProducto).count() == antes_combo
        assert not ctx.report.tenga_errores
    finally:
        db.close()

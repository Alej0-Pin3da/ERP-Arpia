"""Contract tests for migrate.catalog - F0/F1 catalog phase (PR#3 slice).

Covers the STRICT TDD acceptance from tasks #424 T4 (spec R-CAT, design #423):

- Pure: name normalization for dedup, unit resolution, category/unit
  classifier, BOM junk-material filtering, bounded plan building from a
  workbook (insumos + productos + tipos).
- DB (real PostgreSQL via SessionLocal): idempotent upserts for Tipos,
  Insumos (dedup por nombre), Productos (variante NULL / dedup
  por (producto_id, nombre_variante)).
- Integration: dry-run of the real ARPIA.xlsx writes 0 rows (NFR-2); plan
  counts match the workbook; a hand-built mini CatalogPlan commits atomically
  inside session_scope and can be re-applied without duplication.

Test-injected rows use the 'Migratest ' prefix so cleanup never touches
real catalog data; the base tipos/categorias (Lenceria, Corseteria, ...) ARE
the migration's canonical content and may persist across runs (idempotent).
"""

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.db.session import SessionLocal
from app.models import Insumo, Producto, TipoProducto, VarianteProducto
from migrate.context import FaseOptions, MigrationContext

REAL_XLSX = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")

# Names injected by tests; cleanup deletes exactly these, never catalog rows.
PREFIX_TEST = "Migratest"


# --------------------------------------------------------------------------- #
# Fixtures / helpers
# --------------------------------------------------------------------------- #


def _mini_workbook(path: Path) -> None:
    """Mini ARPIA-like workbook: two BOM sheets with junk rows + a mini
    INVENTARIO OCT25 block (cross-sheet dup material)."""
    wb = openpyxl.Workbook()
    bom1 = wb.active
    bom1.title = "CORSET"
    bom1.append(["CORSET", None, None, None, None, None, None, None, "TANGA"])  # R1
    bom1.append(["Producto", "Ancho", "Alto", "cantidad Cms", "valor metro", "valor total"])  # R2
    bom1.append(["Tela Maya Test 1", 64, 37, 2368, 2.5, None])  # R3
    bom1.append(["Encaje Elastico 19 cm negro 10 mts", 28, 18, 504, 1.4, None])  # R4
    bom1.append(["Argolla 90 mm", 2, 1, 2, 72, None])  # R5
    bom1.append(["Horas trabajo", None, None, None, None, None])  # junk
    bom1.append(["COSTO TOTAL CONJUNTO", None, None, None, None, None])  # junk
    bom1.append(["GANANCIA", None, None, None, None, None])  # junk

    bom2 = wb.create_sheet("BLUSAS")
    bom2.append(["MANGA LARGA", None] * 8 + ["MANGA CORTA"])  # R1 title + right block
    bom2.append(["Producto", "Ancho", "Alto", "cantidad Cms", "valor metro", "valor total"])
    bom2.append(["Tela Maya Test 1", 20, 37, 740, 2.5, None])  # dup w/ CORSET
    bom2.append(["Sesgo Elastico 10 mts", 18, 24, 432, 6.3, None])
    bom2.append([4.0, None, None, None, None, None])  # junk numeric

    oct = wb.create_sheet("INVENTARIO OCT25")
    # real sheet (regression against ARPIA.xlsx 2026-08-08): header at R8,
    # data rows R9..29; MATERIAL nombre=B cantidad=D, HERRAJES nombre=F
    # cantidad=H (numero). Columnas A/C/E/G en blanco.
    oct.cell(row=8, column=2, value="MATERIAL")
    oct.cell(row=8, column=4, value="CANTIDAD")
    oct.cell(row=8, column=6, value="HERRAJES")
    oct.cell(row=8, column=8, value="CANTIDAD")
    oct.cell(row=9, column=2, value="Material Migra Test")
    oct.cell(row=9, column=4, value="25 mts")
    oct.cell(row=10, column=6, value="Argolla Migra Test")
    oct.cell(row=10, column=8, value="34")

    # CAJAS combo block: members live in the (nombre, costo, precio) name
    # columns B/F/J (design F3 / bom.py); packaging items (Caja, Vela,
    # Papel, Envio) are insumos too (F1, design D4). Column A does NOT
    # exist on the real sheet, so reading it must not be the F1 path.
    # SHEET_BOUNDS reads CAJAS rows 4..13, so members go on R4+ (col B).
    cajas = wb.create_sheet("CAJAS")
    cajas.cell(row=1, column=2, value="Caja Despertar")  # block header (R1)
    cajas.cell(row=3, column=2, value="Producto")  # header (R3)
    empaques = ["Caja", "Vela", "Papel", "Envio"]
    for i, nombre in enumerate(empaques):
        cajas.cell(row=4 + i, column=2, value=nombre)
    wb.save(path)


@pytest.fixture
def mini_libro(tmp_path) -> Path:
    path = tmp_path / "mini-catalog.xlsx"
    _mini_workbook(path)
    return path


def _borrar_filas_test(db) -> None:
    """Remove rows this test module injected (exact-name matches only)."""
    db.query(Insumo).filter(
        Insumo.nombre.in_(
            [
                "Material Migra Test",
                "Argolla Migra Test",
                "Tela Migra para Upsert",
                f"{PREFIX_TEST} Insumo A",
            ]
        )
    ).delete(synchronize_session=False)
    db.query(Producto).filter(
        Producto.nombre.in_(
            [
                f"{PREFIX_TEST} Corset",
                f"{PREFIX_TEST} Set",
                f"{PREFIX_TEST} P1",
                f"{PREFIX_TEST} P2",
                f"{PREFIX_TEST} Variantes",
                f"{PREFIX_TEST} Precio",
                "Set Celeno",  # producto canonico MIG-2 creado por este modulo
            ]
        )
    ).delete(synchronize_session=False)
    # Remove the canonical catalog tipos that bootstrap_catalogo() inserts.
    # They are migration content, not app seed data; leaving them pollutes the
    # shared per-sheet DB and breaks pagination tests that assume an empty table.
    db.query(TipoProducto).filter(
        TipoProducto.nombre.in_(["Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"])
    ).delete(synchronize_session=False)
    db.commit()


@pytest.fixture(autouse=True, scope="module")
def _cleanup_after_module():
    yield
    db = SessionLocal()
    try:
        _borrar_filas_test(db)
    finally:
        db.close()


@pytest.fixture(scope="module")
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


# --------------------------------------------------------------------------- #
# 1. Pure: normalization / unit / category classification (RED contract)
# --------------------------------------------------------------------------- #


def test_normalizar_nombre_colapsa_espacios_y_case():
    from migrate.catalog import normalizar_nombre

    assert normalizar_nombre("  Encaje    Negro  ") == "Encaje Negro"
    assert normalizar_nombre("café") == "café"  # diacritics kept for display
    assert normalizar_nombre("Tela") == "Tela"  # display keeps case; clave_normalizada folds


def test_clave_normalizada_ignora_acentos_para_dedup():
    from migrate.catalog import clave_normalizada

    assert clave_normalizada("Lencería") == clave_normalizada("Lenceria")
    assert clave_normalizada("Tela Maya") == clave_normalizada(" tela   maya ")


def test_clave_normalizada_colapsa_sinonimos_de_numero():
    """P1 fix: variantes ortograficas de 'numero' se colapsan en la clave.

    'Argolla numero 10 mm' significa 'argolla numero 10' (ordinal), NO un valor
    fijo: la palabra numero/nro/num/No. se ELIMINA cuando precede a un digito,
    unificando la compra 'Argolla numero 10 mm' con el BOM 'Argolla 10 mm'.
    """
    from migrate.catalog import clave_normalizada, normalizar_nombre

    assert clave_normalizada("Argolla numero 10 mm") == clave_normalizada("Argolla 10 mm")
    assert clave_normalizada("Argolla nro 10 mm") == clave_normalizada("Argolla 10 mm")
    assert clave_normalizada("Argolla num 10 mm") == clave_normalizada("Argolla 10 mm")
    assert clave_normalizada("Argolla No. 10 mm") == clave_normalizada("Argolla 10 mm")
    assert clave_normalizada("Cremallera num 3") == clave_normalizada("Cremallera 3")
    # Display name conserva acentos y la palabra original (solo la clave colapsa).
    assert normalizar_nombre("Argolla numero 10 mm") == "Argolla numero 10 mm"
    # Un 'no' dentro de una palabra (p.ej. 'vino') NO se toca.
    assert clave_normalizada("Encaje vino 3") != clave_normalizada("Encaje 3")


def test_resolver_unidad_desde_nombre():
    from migrate.catalog import resolver_unidad

    assert resolver_unidad("Encaje Elastico 19 cm negro 10 mts") == "m"
    assert resolver_unidad("Framilon elastico plano 20 mts") == "m"
    assert resolver_unidad("Sesgo Elastico 10 mts") == "m"
    assert resolver_unidad("Espagueti de pollo 2kg") == "kg"
    assert resolver_unidad("Alambre 500 g") == "g"
    assert resolver_unidad("Sesgo de 2cm") == "cm"
    assert resolver_unidad("Ref 159 24 cm tul bordado rojo pastel") is None  # ancho
    assert resolver_unidad("Tela Maya Ilustrada") is None


def test_clasificar_material_categoria_y_unidad_final():
    from migrate.catalog import clasificar_material

    assert clasificar_material("Ref 159 24 cm tul bordado rojo pastel", None) == ("Telas", "m")
    assert clasificar_material("Argolla 10 mm", None) == ("Herrajes", "un")
    assert clasificar_material("Gafete de 3", None) == ("Empaques", "un")
    assert clasificar_material("Encaje Elastico 19 cm blanco 10 mts", None) == ("Telas", "m")
    assert clasificar_material("Material Migra Test", "25 mts") == ("Telas", "m")
    assert clasificar_material("Argolla Migra Test", "34") == ("Herrajes", "un")


def test_clasificar_barilla_poliester_es_tela_no_herraje():
    """Regression (2026-08, catalogo alineado 16 hojas): 'Barilla poliester
    corset negro 8mm' es un textil continuo que se compra/consume en METROS
    (compra VALQUI '45 mts', BOM 90 cm -> 0.9 m por corset). La keyword
    'barilla' en Herrajes lo clasificaba como Herrajes/un: la compra '45 mts'
    se excluia (EXM-2) y F5 fallaba con InsufficientStockError. Las varillas
    de herraje reales se escriben 'varilla' (Varilla copa brasier talla 30).
    """
    from migrate.catalog import clasificar_material

    assert clasificar_material("Barilla poliester corset negro 8mm", None) == ("Telas", "m")
    # Los herrajes reales con 'varilla' NO cambian de categoria.
    assert clasificar_material("Varilla copa brasier talla 30", None) == ("Herrajes", "un")
    assert clasificar_material("Tapa varilla Negro #1", None) == ("Herrajes", "un")


def test_clasificar_materiales_continuos_por_metro_son_telas():
    """Regression (2026-08): materiales continuos que el workbook compra en
    METROS ('1 mts', '10 mts') y consume en cm en el BOM (Tote Bag usa 48/53
    cm de cadena, 40 cm de cremallera; el corset 30 cm de sesgo rigido).
    Antes caian en Herrajes/un por keywords genericas, la compra se excluia
    (EXM-2) y F5 fallaba con InsufficientStockError. Los herrajes POR PIEZA
    que comparten keyword (deslizadores, ojales, terminales) NO cambian.
    """
    from migrate.catalog import clasificar_material

    assert clasificar_material("Cadena plateada gruesa totebag", None) == ("Telas", "m")
    assert clasificar_material("Cadena gris delgada totebag", None) == ("Telas", "m")
    assert clasificar_material("Cremallera num 3", None) == ("Telas", "m")
    assert clasificar_material("Sesgo rigido para ojales corset", None) == ("Telas", "m")
    # Tapavarilla: nombre '10 mts' explicito -> textil continuo por metro.
    assert clasificar_material("Tapavarilla negro 10 mts", None) == ("Telas", "m")
    # Herrajes por pieza reales (misma keyword) NO cambian.
    assert clasificar_material("deslizadores cremallera num 3", None) == ("Herrajes", "un")
    assert clasificar_material("ojales metalicos 3/8 (grandes)", None) == ("Herrajes", "un")
    assert clasificar_material("Terminales de cordon", None) == ("Herrajes", "un")
    assert clasificar_material("Tapa varilla Negro #1", None) == ("Herrajes", "un")


def test_filtrar_materiales_validos_excluye_junk():
    from migrate.catalog import filtrar_materiales_validos

    filas = [
        {"A": "Tela Maya Test 1"},
        {"A": "Horas trabajo"},
        {"A": "COSTO TOTAL CONJUNTO"},
        {"A": "GANANCIA"},
        {"A": 4.0},
        {"A": "VENTA"},
        {"A": 0},
        {"A": "VLQ"},  # etiqueta de distribucion (Valqui), no material
        {"A": "MAR"},  # etiqueta de distribucion (Margarita)
        {"A": "ARPIA"},  # etiqueta de distribucion
        {"A": "TOTAL BLUSA MANGA CORTA"},  # etiqueta de totales de hoja
        {"I": "Argolla 90 mm"},  # right-block material counts too
    ]
    efectivas = filtrar_materiales_validos(filas)
    assert "Tela Maya Test 1" in efectivas
    assert "Argolla 90 mm" in efectivas
    assert not any(
        m in efectivas
        for m in [
            "Horas trabajo",
            "COSTO TOTAL CONJUNTO",
            "GANANCIA",
            "VENTA",
            "4.0",
            "VLQ",
            "MAR",
            "ARPIA",
            "TOTAL BLUSA MANGA CORTA",
        ]
    )


# --------------------------------------------------------------------------- #
# Plan building from workbook (bounded loader + normalization)
# --------------------------------------------------------------------------- #


def test_plan_workbook_mini_insumos_dedup(mini_libro):
    from migrate.catalog import plan_catalogo
    from migrate.loaders import LibroMigracion

    with LibroMigracion(mini_libro) as libro:
        plan = plan_catalogo(libro)

    # CORSET(4) + BLUSAS adds Sesgo; OCT25 adds 2; CAJAS adds 4 empaques;
    # 'Tela Maya Test 1' dedup across sheets.
    assert plan.conteo_insumos == 10
    nombres = {i.nombre for i in plan.insumos}
    assert {
        "Tela Maya Test 1",
        "Encaje Elastico 19 cm negro 10 mts",
        "Argolla 90 mm",
        "Sesgo Elastico 10 mts",
        "Material Migra Test",
        "Argolla Migra Test",
        "Caja",
        "Vela",
        "Papel",
        "Envio",
    } <= nombres


def test_plan_catalogo_empaques_cajas_entran_al_universo(mini_libro):
    """Regression (2026-08): F1 must read the CAJAS combo member columns
    (B/F/J, design F3 / bom.py) -- NOT column A, which does not exist on the
    sheet -- so the packaging insumos (Caja/Vela/Papel/Envio) enter the F1
    universe. Before this fix F3 silently omitted the 12 combo packaging
    rows because the insumo was never created."""
    from migrate.catalog import plan_catalogo
    from migrate.loaders import LibroMigracion

    with LibroMigracion(mini_libro) as libro:
        plan = plan_catalogo(libro)

    por_categoria: dict[str, list[str]] = {}
    for insumo in plan.insumos:
        por_categoria.setdefault(insumo.categoria, []).append(insumo.nombre)
    empaques = set(por_categoria.get("Empaques", []))
    assert {"Caja", "Vela", "Papel", "Envio"} <= empaques


def test_plan_catalogo_unidades_normalizadas(mini_libro):
    from migrate.catalog import plan_catalogo
    from migrate.loaders import LibroMigracion

    with LibroMigracion(mini_libro) as libro:
        plan = plan_catalogo(libro)

    por = {i.nombre: i for i in plan.insumos}
    assert por["Tela Maya Test 1"].unidad == "m"
    assert por["Tela Maya Test 1"].categoria == "Telas"
    assert por["Argolla 90 mm"].unidad == "un"
    assert por["Argolla 90 mm"].categoria == "Herrajes"
    assert por["Material Migra Test"].unidad == "m"
    assert por["Argolla Migra Test"].unidad == "un"


def test_plan_catalogo_tipos_y_productos_static(mini_libro):
    from migrate.catalog import PRODUCTOS_CATALOGO, TIPOS_CATALOGO, plan_catalogo
    from migrate.loaders import LibroMigracion

    with LibroMigracion(mini_libro) as libro:
        plan = plan_catalogo(libro)

    assert set(plan.tipos) == set(TIPOS_CATALOGO)
    assert plan.conteo_productos == len(PRODUCTOS_CATALOGO)
    for tipo in ["Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"]:
        assert tipo in plan.tipos


# --------------------------------------------------------------------------- #
# DB upserts: idempotencia (real PostgreSQL via SessionLocal)
# --------------------------------------------------------------------------- #


def test_bootstrap_categorias_y_tipos_idempotente(db):
    from migrate.catalog import bootstrap_catalogo

    bootstrap_catalogo(db)
    bootstrap_catalogo(db)

    tipos = {t.nombre for t in db.query(TipoProducto).all()}
    assert {"Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"} <= tipos
    for nombre in ["Lencería", "Corsetería", "Set"]:
        assert db.query(TipoProducto).filter(TipoProducto.nombre == nombre).count() == 1


def test_upsert_insumo_dedup_por_nombre(db):
    from migrate.catalog import bootstrap_catalogo, upsert_insumo

    bootstrap_catalogo(db)
    nombre = "Tela Migra para Upsert"
    a = upsert_insumo(db, nombre, categoria_nombre="Telas")
    b = upsert_insumo(db, nombre, categoria_nombre="Telas")
    assert a.id == b.id
    assert db.query(Insumo).filter(Insumo.nombre == nombre).count() == 1
    fila = db.query(Insumo).filter(Insumo.nombre == nombre).first()
    assert fila.stock_actual == 0 and fila.stock_minimo == 0 and fila.costo_promedio_actual == 0
    assert fila.categoria.nombre == "Telas"


def test_upsert_producto_sin_variante_no_duplica(db):
    from migrate.catalog import bootstrap_catalogo, upsert_producto

    bootstrap_catalogo(db)
    nombre = f"{PREFIX_TEST} Corset"
    a = upsert_producto(db, nombre, tipo="Corsetería")
    b = upsert_producto(db, nombre, tipo="Corsetería")
    assert a.id == b.id
    assert db.query(Producto).filter(Producto.nombre == nombre).count() == 1
    assert len(a.variantes) == 0  # variante NULL: sin fila duplicada vacia


def test_upsert_producto_variantes_dedup(db):
    from migrate.catalog import bootstrap_catalogo, upsert_producto

    bootstrap_catalogo(db)
    nombre = f"{PREFIX_TEST} Set"
    a = upsert_producto(db, nombre, tipo="Set", variantes=["S", "S", "M"])
    assert {v.nombre_variante for v in a.variantes} == {"S", "M"}
    b = upsert_producto(db, nombre, tipo="Set", variantes=["S", "M"])
    assert b.id == a.id
    assert len(b.variantes) == 2  # dedup por (producto_id, nombre_variante)


# --------------------------------------------------------------------------- #
# Phase level: dry-run plan on real workbook + atomic mini apply
# --------------------------------------------------------------------------- #


def test_catalogar_dry_run_real_no_escribe():
    from migrate.catalog import PRODUCTOS_CATALOGO, catalogar

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    db = SessionLocal()
    try:
        antes = (
            db.query(TipoProducto).count(),
            db.query(Insumo).count(),
            db.query(Producto).count(),
        )
        ctx = MigrationContext.para_fase(FaseOptions(source=REAL_XLSX, modo="dry-run"), "F1")
        plan = catalogar(ctx)
        despues = (
            db.query(TipoProducto).count(),
            db.query(Insumo).count(),
            db.query(Producto).count(),
        )
        # NFR-2: dry-run termina con 0 filas escritas
        assert antes == despues
        assert plan.conteo_insumos >= 20  # recetas BOM + herrajes reales
        assert plan.conteo_productos == len(PRODUCTOS_CATALOGO)
    finally:
        db.close()


def test_oct25_layout_real_entra_al_universo_f1():
    """Regression: INVENTARIO OCT25 uses MATERIAL B/D + HERRAJES F/H (real
    ARPIA.xlsx layout, verified 2026-08-08). The 20 real materials (B) and the
    14 real herrajes (F) MUST enter the F1 insumo universe, and the quantity
    cells ('9,5 mts', '11 mts'...) MUST NOT be treated as material names."""
    from migrate.catalog import _leer_materiales, clave_normalizada
    from migrate.loaders import LibroMigracion

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    materiales_b = [
        "Encaje negro sin pelitos",
        "Tela entrepierna negra",
        "Tela entrepierna blanca",
        "Encaje blanco chantilli (pelitos) para bicolor",
        "Encaje negro chantilli (pelitos) para bicolor",
        "Tira de brasier blanca",
        "Contorno para Bustier negro 2 cm ancho",
        "Tapa varilla Negro #1",
        "Tapa varilla Negro #2",
        "Elastico de contorno de 1 cm blanco",
        "Tapa Costura Negro",
        "Elastico plano negro",
        "Elastico plano blanco",
        "Tira de brasier negra",
        "Sesgo de 2cm blanco",
        "Sesgo de 2cm negro",
        "Mallatex negra",
        "Mallatex blanca",
        "Ref 100 24 cm tul bordado negro",
        "Ref 159 24 cm tul bordado rojo pastel",
    ]
    herrajes = [
        "Argollas grandes",
        "* Argollas Medianas",
        "* Argollas Pequenas",
        "* Ochos Grandes",
        "* Ochos Medianos",
        "* Ochos Pequenos",
        "* Gancho G grandes",
        "* Gancho G Medianos",
        "* Ganchos G Pequenos",
        "Varilla copa brasier talla 30",
        "Varilla copa brasier talla 32",
        "Varilla copa brasier talla 36",
        "Varilla copa brasier talla 34",
        "Variila plastica cortada 18cms",
    ]
    with LibroMigracion(REAL_XLSX) as libro:
        universo = _leer_materiales(libro, None)
    claves = set(universo)
    for nombre in materiales_b + herrajes:
        assert clave_normalizada(nombre) in claves, f"material OCT25 ausente: {nombre}"
    # Cantidades de D (9,5 mts / 11 mts) nunca son nombres de insumo.
    assert clave_normalizada("9,5 mts") not in claves
    assert clave_normalizada("11 mts") not in claves


def test_aplicar_plan_transaccional_y_cleanup(db):
    from migrate.catalog import (
        CatalogPlan,
        InsumoPlan,
        ProductoPlan,
        aplicar_plan,
    )

    plan = CatalogPlan(
        tipos=["Lencería"],
        insumos=[InsumoPlan(nombre=f"{PREFIX_TEST} Insumo A", unidad="m", categoria="Telas")],
        productos=[
            ProductoPlan(nombre=f"{PREFIX_TEST} P1", tipo="Set"),
            ProductoPlan(nombre=f"{PREFIX_TEST} P2", tipo="Set", variantes=("XS", "S")),
        ],
    )
    aplicar_plan(db, plan)
    db.commit()

    assert db.query(Producto).filter(Producto.nombre == f"{PREFIX_TEST} P2").count() == 1
    assert db.query(Insumo).filter(Insumo.nombre == f"{PREFIX_TEST} Insumo A").count() == 1


# --------------------------------------------------------------------------- #
# Variantes XXS-XL + Set Celeno (MIG-1/MIG-2, tallas desde la XXS hasta la XL)
# --------------------------------------------------------------------------- #

TALLAS_XXS_XL = ("XXS", "XS", "S", "M", "L", "XL")


def test_plan_catalogo_real_30_variantes_y_14_productos():
    """MIG-1/MIG-2: 5 productos tallados x 6 tallas = 30 variantes; el catalogo
    crece a 14 con Set Celeno @ 75000 (locked decision, precio_venta NULL en
    las variantes porque comparten el precio del producto)."""
    from migrate.catalog import plan_catalogo
    from migrate.loaders import LibroMigracion

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    with LibroMigracion(REAL_XLSX) as libro:
        plan = plan_catalogo(libro)

    assert plan.conteo_productos == 14
    por_nombre = {p.nombre: p for p in plan.productos}
    for tallado in (
        "Set Aelo",
        "Set Ocipete",
        "Set Celeno",
        "Blusa Manga Larga",
        "Blusa Manga Corta",
    ):
        assert por_nombre[tallado].variantes == TALLAS_XXS_XL, tallado
    assert sum(len(p.variantes) for p in plan.productos) == 30
    celeno = por_nombre["Set Celeno"]
    assert celeno.precio_sugerido == Decimal("75000")


def test_plan_catalogo_real_sin_variantes_en_corset_garras_y_combos():
    """MIG-1: un producto sin tupla variantes (Corset Garras, combos) NUNCA
    recibe filas de variante (tampoco una variante NULL fantasma)."""
    from migrate.catalog import plan_catalogo
    from migrate.loaders import LibroMigracion

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    with LibroMigracion(REAL_XLSX) as libro:
        plan = plan_catalogo(libro)
    por_nombre = {p.nombre: p for p in plan.productos}
    for sin_tallas in ("Corset Garras", "Caja Despertar", "Caja Saca Las Garras"):
        assert por_nombre[sin_tallas].variantes == ()


def test_aplicar_plan_set_celeno_precio_y_reapply(db):
    """MIG-2: aplicar_plan persiste Set Celeno @ 75000 y un re-apply lo
    mantiene (D1: el plan/catalogo es la fuente de verdad del precio)."""
    from migrate.catalog import CatalogPlan, ProductoPlan, aplicar_plan

    plan = CatalogPlan(
        tipos=["Set"],
        productos=[
            ProductoPlan(
                nombre="Set Celeno",
                tipo="Set",
                variantes=TALLAS_XXS_XL,
                precio_sugerido=Decimal("75000"),
            )
        ],
    )
    aplicar_plan(db, plan)
    db.commit()
    p = db.query(Producto).filter(Producto.nombre == "Set Celeno").one()
    assert p.precio_venta_sugerido == Decimal("75000")
    assert len(p.variantes) == 6

    aplicar_plan(db, plan)  # re-run: idempotente, precio estable
    db.commit()
    db.refresh(p)
    assert p.precio_venta_sugerido == Decimal("75000")
    assert db.query(VarianteProducto).filter(VarianteProducto.producto_id == p.id).count() == 6


def test_upsert_producto_variantes_duplicadas_6_filas(db):
    """MIG-1: dedup por (producto_id, nombre_variante): re-upsert de las 6
    tallas NO duplica filas (guard manual antes del UNIQUE)."""
    from migrate.catalog import bootstrap_catalogo, upsert_producto

    bootstrap_catalogo(db)
    nombre = f"{PREFIX_TEST} Variantes"
    a = upsert_producto(db, nombre, tipo="Set", variantes=TALLAS_XXS_XL)
    b = upsert_producto(db, nombre, tipo="Set", variantes=TALLAS_XXS_XL)
    assert b.id == a.id
    assert {v.nombre_variante for v in a.variantes} == set(TALLAS_XXS_XL)
    assert db.query(VarianteProducto).filter(VarianteProducto.producto_id == a.id).count() == 6


def test_upsert_producto_refresca_precio_solo_cuando_no_none(db):
    """D1: upsert_producto refresca precio_venta_sugerido en el producto
    existente SOLO cuando precio_sugerido != None; un caller viejo que no
    pasa precio nunca pisa el valor del catalogo."""
    from migrate.catalog import bootstrap_catalogo, upsert_producto

    bootstrap_catalogo(db)
    nombre = f"{PREFIX_TEST} Precio"
    p1 = upsert_producto(db, nombre, tipo="Accesorio")
    assert p1.precio_venta_sugerido == Decimal("0")
    p2 = upsert_producto(db, nombre, tipo="Accesorio", precio_sugerido=Decimal("75000"))
    assert p2.id == p1.id
    assert p2.precio_venta_sugerido == Decimal("75000")
    p3 = upsert_producto(db, nombre, tipo="Accesorio")  # sin precio: no pisa
    assert p3.precio_venta_sugerido == Decimal("75000")

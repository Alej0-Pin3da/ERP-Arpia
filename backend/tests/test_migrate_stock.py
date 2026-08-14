"""Contract tests for migrate.stock - F4 stock inicial OCT25 (PR#5 slice).

Covers STRICT TDD acceptance from tasks #424 T7 (spec EXM-2/3/4, NFR-1/2;
design #423 stock.py + D4):

- Quantity: the OCT25 cell may be a bare number (150) -> taken as-is in the
  insumo canonical unit; or a quantity+unit string ('11 mts', '50 cm')
  converted to the canonical unit of the insumo (same normalizer as F2,
  EXM-2). Uninterpretable values (None, #DIV/0!, junk) -> None (reported,
  never inferred).
- Workbook -> plan: INVENTARIO OCT25 MATERIAL block (A=nombre, B=cantidad)
  and HERRAJES block (D=nombre, E=cantidad); junk/totals excluded (the
  PRENDAS block in G is not read).
- DB (real PostgreSQL): applying sets stock_actual to the OCT25 snapshot ONLY
  (the cost stays as the F2 WAC left it: insumo con WAC -> costo preservado;
  insumo sin WAC -> 0). No UNIQUE on stock: idempotence = re-setting the
  same value, never adding or duplicating rows.
- Phase: F4 registered in the runner registry; dry-run on the real workbook
  writes 0 rows (NFR-2).

Test-injected rows use the 'Migratest ' prefix so cleanup never touches real
migration data; the canonical catalog tipos inserted by bootstrap_catalogo()
are removed at module cleanup (same pattern as the other test_migrate_*).
"""

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.db.session import SessionLocal
from app.models import CompraInsumo, Insumo, TipoProducto
from migrate.context import FaseOptions, MigrationContext
from migrate.loaders import LibroMigracion

REAL_XLSX = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")

# Names injected by tests; cleanup deletes exactly these, never catalog rows.
P = "Migratest"
P_TELA = f"{P} Stock Tela"
P_ARG = f"{P} Stock Argolla"
P_LINO = f"{P} Stock Lino"


# --------------------------------------------------------------------------- #
# Mini workbook builder (real layout: header R8, data R9+ per SHEET_BOUNDS)
# --------------------------------------------------------------------------- #


def _mini_workbook(path: Path) -> None:
    """Mini OCT25: MATERIAL (B/D), HERRAJES (F/H), PRENDAS (J..O) + junk."""
    wb = openpyxl.Workbook()
    oct25 = wb.active
    oct25.title = "INVENTARIO OCT25"
    # R8 header; data R9+ (SHEET_BOUNDS 9..29). Real layout, verified 2026-08-08.
    oct25.cell(row=8, column=2, value="MATERIAL")  # B8
    oct25.cell(row=8, column=4, value="CANTIDAD")  # D8
    oct25.cell(row=8, column=6, value="HERRAJES")  # F8
    oct25.cell(row=8, column=8, value="CANTIDAD")  # H8
    oct25.cell(row=8, column=10, value="PRENDAS")  # J8
    oct25.cell(row=8, column=13, value="PRECIO VENTA")  # M8
    # Data rows R9+ (within SHEET_BOUNDS 9..29).
    oct25.cell(row=9, column=2, value=P_TELA)  # B9 material
    oct25.cell(row=9, column=4, value="11 mts")  # D9 cantidad
    oct25.cell(row=9, column=6, value=P_ARG)  # F9 herraje
    oct25.cell(row=9, column=8, value="150")  # H9 cantidad
    oct25.cell(row=9, column=10, value="Corset")  # J9 PRENDAS: never read
    oct25.cell(row=9, column=13, value=75000)  # M9 precio (reference)
    oct25.cell(row=10, column=2, value=P_LINO)
    oct25.cell(row=10, column=4, value="50 cm")
    oct25.cell(row=10, column=6, value="Sublimacion (cm2)")
    oct25.cell(row=10, column=8, value="4670")
    oct25.cell(row=11, column=2, value="GANANCIA")  # junk (subcadena)
    oct25.cell(row=11, column=6, value="TOTAL HERR")  # junk
    oct25.cell(row=12, column=2, value=4.0)  # junk numeric
    wb.save(path)


def _mini_workbook_una_fila(
    path: Path, col_nombre: int, nombre: str, col_cant: int, cantidad: object
) -> None:
    """Mini OCT25 con UNA fila de datos (header row 8, data row 9)."""
    wb = openpyxl.Workbook()
    oct25 = wb.active
    oct25.title = "INVENTARIO OCT25"
    oct25.cell(row=8, column=2, value="MATERIAL")  # B8
    oct25.cell(row=8, column=4, value="CANTIDAD")  # D8
    oct25.cell(row=8, column=6, value="HERRAJES")  # F8
    oct25.cell(row=8, column=8, value="CANTIDAD")  # H8
    oct25.cell(row=9, column=col_nombre, value=nombre)
    oct25.cell(row=9, column=col_cant, value=cantidad)
    wb.save(path)


@pytest.fixture
def mini_stock(tmp_path) -> Path:
    path = tmp_path / "mini-oct25.xlsx"
    _mini_workbook(path)
    return path


# --------------------------------------------------------------------------- #
# Module-level DB cleanup (canonical tipos + test rows)
# --------------------------------------------------------------------------- #


def _borrar_filas_test(db) -> None:
    """Remove rows this test module injected (exact-name matches only)."""
    insumos = db.query(Insumo).filter(Insumo.nombre.in_([P_TELA, P_ARG, P_LINO])).all()
    for ins in insumos:
        # registrar_compra lega CompraInsumo (FK al insumo); borrar primero.
        db.query(CompraInsumo).filter(CompraInsumo.insumo_id == ins.id).delete(
            synchronize_session=False
        )
    db.query(Insumo).filter(Insumo.nombre.in_([P_TELA, P_ARG, P_LINO])).delete(
        synchronize_session=False
    )
    # Remove the canonical catalog tipos that bootstrap_catalogo() inserts.
    db.query(TipoProducto).filter(
        TipoProducto.nombre.in_(["Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"])
    ).delete(synchronize_session=False)
    db.commit()


@pytest.fixture(autouse=True, scope="module")
def _cleanup_after_module():
    yield
    db = SessionLocal()
    try:
        _borrar_filas_test(db)
    finally:
        db.close()


@pytest.fixture(scope="module")
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _preparar_insumos(db) -> None:
    """F1-ish: bootstrap + 3 insumos de prueba (stock=0, costo=0 explicitos)."""
    from migrate.catalog import bootstrap_catalogo, upsert_insumo

    bootstrap_catalogo(db)
    upsert_insumo(db, P_TELA, categoria_nombre="Telas")
    upsert_insumo(db, P_ARG, categoria_nombre="Herrajes")
    upsert_insumo(db, P_LINO, categoria_nombre="Telas")
    db.commit()


# --------------------------------------------------------------------------- #
# 1. Pure: cantidad OCT25 -> unidad canonica (EXM-2, reusa normalize)
# --------------------------------------------------------------------------- #


def test_cantidad_stock_parsea_unidad_tal_cual():
    from migrate.stock import normalizar_cantidad_stock

    # '11 mts' en la unidad canonica de Telas (m) -> 11
    assert normalizar_cantidad_stock("11 mts", "m") == Decimal("11")
    # '150' en Herrajes (un) -> 150 piezas
    assert normalizar_cantidad_stock("150", "un") == Decimal("150")
    # '50 cm' en Telas -> 0.5 m (conversion lineal, EXM-2)
    assert normalizar_cantidad_stock("50 cm", "m") == Decimal("0.5")


def test_cantidad_stock_rechaza_no_interpretable():
    from migrate.stock import normalizar_cantidad_stock

    assert normalizar_cantidad_stock(None, "m") is None
    assert normalizar_cantidad_stock("#DIV/0!", "m") is None
    assert normalizar_cantidad_stock("", "m") is None
    assert normalizar_cantidad_stock("0", "m") is None  # sin stock real


# --------------------------------------------------------------------------- #
# 2. Workbook -> plan (bloques MATERIAL/HERRAJES, junk excluido)
# --------------------------------------------------------------------------- #


def test_plan_stock_lee_bloques_material_y_herrajes(mini_stock):
    from migrate.stock import plan_stock

    with LibroMigracion(mini_stock) as libro:
        plan = plan_stock(libro)

    por = {s.insumo_nombre: s for s in plan.stock}
    assert P_TELA in por and por[P_TELA].cantidad == Decimal("11")
    assert P_ARG in por and por[P_ARG].cantidad == Decimal("150")
    assert P_LINO in por and por[P_LINO].cantidad == Decimal("0.5")


def test_plan_stock_excluye_junk_y_prendas(mini_stock):
    from migrate.stock import plan_stock

    with LibroMigracion(mini_stock) as libro:
        plan = plan_stock(libro)

    nombres = {s.insumo_nombre for s in plan.stock}
    assert "GANANCIA" not in nombres
    assert "TOTAL HERR" not in nombres
    assert "Corset" not in nombres  # bloque PRENDAS no se lee
    assert all(s.cantidad > 0 for s in plan.stock)


# --------------------------------------------------------------------------- #
# 3. DB: aplicar stock (snapshot) + idempotencia + costo WAC intacto
# --------------------------------------------------------------------------- #


def test_aplicar_stock_setea_stock_actual_y_deja_costo(db, mini_stock):
    from migrate.stock import aplicar_stock, plan_stock

    _preparar_insumos(db)
    with LibroMigracion(mini_stock) as libro:
        plan = plan_stock(libro)
    aplicar_stock(db, plan)
    db.commit()

    ins = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    assert ins.stock_actual == Decimal("11")
    assert ins.stock_minimo == Decimal("0")
    assert ins.costo_promedio_actual == Decimal("0")  # sin WAC -> 0


def test_aplicar_stock_idempotente(db, mini_stock):
    from migrate.stock import aplicar_stock, plan_stock

    _preparar_insumos(db)
    with LibroMigracion(mini_stock) as libro:
        plan = plan_stock(libro)
    aplicar_stock(db, plan)
    db.commit()
    aplicar_stock(db, plan)
    db.commit()

    ins = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    assert ins.stock_actual == Decimal("11")  # no se suma, no se duplica


def test_aplicar_stock_no_pisa_costo_wac(db, mini_stock):
    from app.services.wac import registrar_compra
    from migrate.stock import aplicar_stock, plan_stock

    _preparar_insumos(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).one()
    # Estado limpio (los tests comparten la sesion de modulo): stock 0 -> WAC
    # deja costo 100; luego el snapshot OCT25 setea stock sin tocar el costo.
    tela.stock_actual = Decimal("0")
    tela.costo_promedio_actual = Decimal("0")
    db.flush()
    registrar_compra(db, tela.id, None, Decimal("5"), Decimal("100"), commit=False)
    db.commit()
    assert tela.costo_promedio_actual == Decimal("100")  # WAC de F2 dejado

    with LibroMigracion(mini_stock) as libro:
        plan = plan_stock(libro)
    aplicar_stock(db, plan)
    db.commit()

    db.refresh(tela)
    assert tela.stock_actual == Decimal("11")  # snapshot OCT25
    assert tela.costo_promedio_actual == Decimal("100")  # WAC preservado


# --------------------------------------------------------------------------- #
# 3b. Aliases OCT25 -> catalogo (ALIASES_STOCK_A_CATALOGO, patron de bom.py)
# --------------------------------------------------------------------------- #


def test_aplicar_stock_alias_tira_brasier_blanca(db, tmp_path):
    from migrate.catalog import upsert_insumo
    from migrate.stock import aplicar_stock, plan_stock

    _preparar_insumos(db)
    canonico = upsert_insumo(db, "Tira de Brasier blanco 10 mts")
    db.commit()

    path = tmp_path / "alias-tira.xlsx"
    _mini_workbook_una_fila(path, 2, "Tira de brasier blanca", 4, "7 mts")
    with LibroMigracion(path) as libro:
        plan = plan_stock(libro)
    res = aplicar_stock(db, plan)
    db.commit()

    assert res["seteados"] == 1
    assert res["omitidos"] == 0
    db.refresh(canonico)
    assert canonico.stock_actual == Decimal("7")
    # El alias NUNCA crea insumos: el nombre corto no existe en el catalogo.
    assert db.query(Insumo).filter(Insumo.nombre == "Tira de brasier blanca").count() == 0


def test_aplicar_stock_alias_argollas_medianas_estrella(db, tmp_path):
    from migrate.catalog import upsert_insumo
    from migrate.stock import aplicar_stock, plan_stock

    _preparar_insumos(db)
    canonico = upsert_insumo(db, "Argolla numero 8 mm")
    db.commit()

    path = tmp_path / "alias-argollas.xlsx"
    _mini_workbook_una_fila(path, 6, " * Argollas Medianas ", 8, 120)
    with LibroMigracion(path) as libro:
        plan = plan_stock(libro)
    res = aplicar_stock(db, plan)
    db.commit()

    assert res["seteados"] == 1
    assert res["omitidos"] == 0
    db.refresh(canonico)
    assert canonico.stock_actual == Decimal("120")


def test_aplicar_stock_alias_varilla_copa_talla_34(db, tmp_path):
    from migrate.catalog import upsert_insumo
    from migrate.stock import aplicar_stock, plan_stock

    _preparar_insumos(db)
    canonico = upsert_insumo(db, "ARCO METALICO 2001 34")
    db.commit()

    path = tmp_path / "alias-varilla.xlsx"
    _mini_workbook_una_fila(path, 6, "Varilla copa brasier talla 34", 8, 50)
    with LibroMigracion(path) as libro:
        plan = plan_stock(libro)
    res = aplicar_stock(db, plan)
    db.commit()

    assert res["seteados"] == 1
    assert res["omitidos"] == 0
    db.refresh(canonico)
    assert canonico.stock_actual == Decimal("50")


def test_aplicar_stock_sin_alias_ni_match_exacto_se_omite(db, tmp_path):
    from migrate.stock import aplicar_stock, plan_stock

    _preparar_insumos(db)
    nombre_faltante = "Migratest Sin Alias"
    path = tmp_path / "sin-alias.xlsx"
    _mini_workbook_una_fila(path, 2, nombre_faltante, 4, "3 mts")
    with LibroMigracion(path) as libro:
        plan = plan_stock(libro)
    res = aplicar_stock(db, plan)
    db.commit()

    assert res["omitidos"] == 1
    assert res["seteados"] == 0
    assert db.query(Insumo).filter(Insumo.nombre == nombre_faltante).count() == 0


# --------------------------------------------------------------------------- #
# 4. Phase: runner registry + real workbook dry run (NFR-2)
# --------------------------------------------------------------------------- #


def test_f4_registrada_en_runner():
    from migrate import FASE_RUNNERS, FASES, FASES_IMPLEMENTADAS

    assert any(f.id == "F4" for f in FASES)
    assert "F4" in FASE_RUNNERS
    assert "F4" in FASES_IMPLEMENTADAS


def test_cargar_stock_dry_run_real_no_escribe():
    from migrate.stock import cargar_stock

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    db = SessionLocal()
    try:
        antes = db.query(Insumo).count()
        ctx = MigrationContext.para_fase(FaseOptions(source=REAL_XLSX, modo="dry-run"), "F4")
        cargar_stock(ctx)
        assert db.query(Insumo).count() == antes
        assert not ctx.report.tenga_errores
    finally:
        db.close()

"""Contract tests for migrate.validate - F7 validacion transversal (PR#8).

Covers STRICT TDD acceptance from tasks #424 T10 / design #423 validate.py +
checks_n7 (spec EXM-5, NFR-1/2; user slice-8 contract N7a-g):

- N7a "conteos": rows per domain (tipos, proveedores, insumos, productos,
  BOM_-insumos, compras, ventas, movimientos, socios) counted in the DB
  against the plan's expected rows (built from the source workbook's own
  plans). estado: OK flat; WARN when the DB has FEWER rows than the plan
  (missing / omitted with cause); ERROR when more (a re-run duplicated).
- N7b "stock_negativo": no Insumo with stock_actual < 0 (EXM-5).
- N7c "finanzas": every MovimientoFinanciero has monto > 0 and the sum per
  tipo matches the source plan; socios sum == 100 (FIN-2).
- N7d "cuadre": for each snapshot insumo, stock_actual == snapshot + compras
  - explosion of the ventas (VTA-3), tolerance 0.0005. A snapshot insumo
  absent from the catalog breaks the precondition -> ERROR; divergence with
  a cause -> WARN; exact match -> PASS.
- N7e "precios": every DetalleVenta costo_unitario_aplicado == the historical
  cost of the source row (VTA-1/EXM-5e) and Venta.total_venta ==
  quantity*precio (no double discount, VTA-2). Divergence -> ERROR.
- N7f "fechas": no migration row carries a fake now() date (D5): compras,
  ventas and movimientos are all far from the server now.
- N7g "idempotencia": natural keys of compras/movimientos/BOM-Insumos with
  count > 1 -> ERROR; ventas count per key > plan count -> ERROR.

Test-injected rows use the 'Migratest ' prefix so cleanup never touches real
migration data; the canonical catalog tipos inserted by bootstrap_catalogo()
are removed at module cleanup (same pattern as the other test_migrate_*).
"""

from datetime import datetime, timezone, timedelta
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app.db.session import SessionLocal
from app.models import (
    BomInsumo,
    Cliente,
    CompraInsumo,
    DetalleVenta,
    Insumo,
    MovimientoFinanciero,
    Producto,
    Proveedor,
    SociosConfiguracion,
    TipoProducto,
    Venta,
)
from migrate.context import FaseOptions, MigrationContext
from migrate.loaders import LibroMigracion

REAL_XLSX = Path(r"C:\wamp64\www\ERP-Arpia\ARPIA.xlsx")
PREFIX = "Migratest"
P = f"{PREFIX} Validate"

# Names injected by tests; cleanup deletes exactly these, never catalog rows.
P_TELA = f"{P} Tela"
P_PROV = f"{P} Proveedor"
# Producto de prueba UNICO (no 'Corset'): la DB real ya tiene el producto
# 'Corset' de la migracion cargada con combos BOM_Productos que lo referencian
# (FK RESTRICT); usar ese nombre romperia el cleanup por FK violation.
P_PROD = f"{P} Corset"
P_VAR = "S"
P_CLI = f"{P} Cliente"
P_MOV = f"{P} Movimiento"

FECHA_VENTA = datetime(2024, 10, 20, tzinfo=timezone.utc)
FECHA_VENTA2 = datetime(2024, 11, 5, tzinfo=timezone.utc)
FECHA_COMPRA = datetime(2025, 9, 15, tzinfo=timezone.utc)
FECHA_COMPRA2 = datetime(2025, 9, 20, tzinfo=timezone.utc)
FECHA_MOV = datetime(2025, 8, 1, tzinfo=timezone.utc)

PRECIO_VENTA = Decimal("71250")
COSTO_VENTA = Decimal("26109")
TOLERANCIA = Decimal("0.001")

# The 3 canonical partners of the migration (spec FIN-2): sum==100.
SOCIOS_ESPERADOS = {
    "Valqui": Decimal("40"),
    "Margarita": Decimal("30"),
    "ARPIA": Decimal("30"),
}

P_EXTRA = f"{P} Extra"  # snapshot insumo OUTSIDE the catalog (N7d precondicion)


# --------------------------------------------------------------------------- #
# Mini workbook builder (real sheet layouts consumed by the phase plans)
# --------------------------------------------------------------------------- #


def _mini_workbook(path: Path, extra_oct25: tuple[str, str] | None = None) -> None:
    """Mini validate workbook.

    - CORSET sheet: 1 material (the BOM universe for F2/F3 catalog).
    - INVENTARIO OCT25: snapshot stock for that material (10 mts).
    - INVERSION VALQUI: a BOM purchase row (-> F2 WAC) + an equipment row
      (non-BOM -> F6 movement).
    - VENTAS: one historical sale of the product 'Corset'.
    - Proveedores: one supplier.

    ``extra_oct25`` adds a HERRAJES snapshot row (name, quantity) whose insumo
    is NOT created in the catalog (used to break the N7d precondition).
    """
    wb = openpyxl.Workbook()

    prov = wb.active
    prov.title = "Proveedores"
    prov.append(["Proveedor", "URL", "Precio Unidad", "Ubicacion", "Contactado"])
    prov.append(["", P_PROV, None, "Cali", "SI"])  # B=nombre

    bom = wb.create_sheet("CORSET")
    bom.append(["CORSET", None, None, None, None, None, None, None, "TANGA"])
    bom.append(["Producto", "Ancho", "Alto", "cantidad Cms", "valor metro", "valor total"])
    bom.append([P_TELA, 64, 37, 2368, 2.5, None])

    oct25 = wb.create_sheet("INVENTARIO OCT25")
    oct25.cell(row=8, column=2, value="MATERIAL")
    oct25.cell(row=8, column=4, value="CANTIDAD")
    oct25.cell(row=8, column=6, value="HERRAJES")
    oct25.cell(row=8, column=8, value="CANTIDAD")
    oct25.cell(row=9, column=2, value=P_TELA)
    oct25.cell(row=9, column=4, value="10 mts")
    if extra_oct25 is not None:
        nombre_extra, cant_extra = extra_oct25
        oct25.cell(row=9, column=6, value=nombre_extra)  # F: HERRAJES nombre
        oct25.cell(row=9, column=8, value=cant_extra)    # H: HERRAJES cantidad

    ventas = wb.create_sheet("VENTAS")
    ventas.append(["Producto", None, None, None, None, None, "Precio Venta",
                   "Costo", "Ganancias", None, None, None, "Fecha", "Color",
                   "Desc", "Cliente"])
    ventas.append([P_PROD, P_VAR, None, None, None, None, 71250.0, COSTO_VENTA,
                   45141, None, None, None, datetime(2024, 10, 20), "vino",
                   None, P_CLI])

    inv = wb.create_sheet("INVERSION VALQUI")
    inv.cell(row=2, column=1, value="Cantidad")
    inv.cell(row=2, column=2, value="Producto")
    inv.cell(row=2, column=4, value="Costo")
    inv.cell(row=2, column=5, value="Fecha")
    inv.cell(row=2, column=6, value="Provedor")
    # R3: BOM purchase (tela) -> compra WAC F2.
    inv.cell(row=3, column=1, value="2 mts")
    inv.cell(row=3, column=2, value=P_TELA)
    inv.cell(row=3, column=4, value=200)
    inv.cell(row=3, column=5, value=datetime(2025, 9, 15))
    inv.cell(row=3, column=6, value=P_PROV)
    # R4: non-BOM equipment -> movimiento F6.
    inv.cell(row=4, column=1, value=1)
    inv.cell(row=4, column=2, value=P_MOV)
    inv.cell(row=4, column=4, value=320000)
    inv.cell(row=4, column=5, value=datetime(2025, 8, 1))
    wb.save(path)


@pytest.fixture
def mini_libro(tmp_path) -> Path:
    path = tmp_path / "mini-validate.xlsx"
    _mini_workbook(path)
    return path


# --------------------------------------------------------------------------- #
# Module-level DB cleanup (canonical tipos + test rows + socios)
# --------------------------------------------------------------------------- #


def _borrar_detalles_ventas_p(db) -> None:
    """Quita Detalle_Ventas/Ventas que referencian el producto test."""
    vids = set(
        db.query(DetalleVenta.venta_id)
        .filter(DetalleVenta.producto_id.in_(
            db.query(Producto.id).filter(Producto.nombre == P_PROD)
        ))
        .all()
    )
    vids |= set(
        db.query(Venta.id)
        .join(Cliente, Venta.cliente_id == Cliente.id)
        .filter(Cliente.nombre == P_CLI)
        .all()
    )
    for (vid,) in vids:
        db.query(DetalleVenta).filter(DetalleVenta.venta_id == vid).delete(
            synchronize_session=False
        )
        db.query(Venta).filter(Venta.id == vid).delete(
            synchronize_session=False
        )
    db.query(Cliente).filter(Cliente.nombre == P_CLI).delete(
        synchronize_session=False
    )


def _borrar_test(db) -> None:
    """Borra SOLO filas de test por nombre exacto/parejas exactas."""
    _borrar_detalles_ventas_p(db)
    db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.descripcion == P_MOV
    ).delete(synchronize_session=False)
    # Compras y BOM que referencian al insumo test (FK RESTRICT).
    db.query(CompraInsumo).filter(
        CompraInsumo.insumo_id.in_(
            db.query(Insumo.id).filter(Insumo.nombre == P_TELA)
        )
    ).delete(synchronize_session=False)
    db.query(BomInsumo).filter(
        BomInsumo.insumo_id.in_(
            db.query(Insumo.id).filter(Insumo.nombre == P_TELA)
        )
    ).delete(synchronize_session=False)
    db.query(BomInsumo).filter(
        BomInsumo.producto_id.in_(
            db.query(Producto.id).filter(Producto.nombre == P_PROD)
        )
    ).delete(synchronize_session=False)
    db.query(Insumo).filter(Insumo.nombre == P_TELA).delete(
        synchronize_session=False
    )
    db.query(Producto).filter(Producto.nombre == P_PROD).delete(
        synchronize_session=False
    )
    db.query(Proveedor).filter(Proveedor.nombre == P_PROV).delete(
        synchronize_session=False
    )
    db.commit()


def _borrar_socios_y_tipos(db) -> None:
    for nombre in SOCIOS_ESPERADOS:
        socio = db.query(SociosConfiguracion).filter(
            SociosConfiguracion.nombre == nombre
        ).first()
        if socio is None:
            continue
        con_movimientos = db.query(MovimientoFinanciero).filter(
            MovimientoFinanciero.socio_id == socio.id
        ).first()
        if con_movimientos is None:
            db.delete(socio)
    # Tipos canonicos: se borran SOLO si ningun producto (real de la migracion
    # o de otro modulo) los referencia; con la migracion cargada los productos
    # reales usan estos tipos -> se conservan (patron _borrar_filas_test).
    for nombre_tipo in ["Lencería", "Corsetería", "Blusa", "Accesorio", "Set", "Combo"]:
        tipo = db.query(TipoProducto).filter(TipoProducto.nombre == nombre_tipo).first()
        if tipo is None:
            continue
        con_productos = db.query(Producto).filter(
            Producto.tipo_producto_id == tipo.id
        ).first()
        if con_productos is None:
            db.delete(tipo)
    db.commit()


@pytest.fixture(scope="module")
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


@pytest.fixture(autouse=True, scope="module")
def _cleanup_after_module(db):
    yield
    _borrar_test(db)
    _borrar_socios_y_tipos(db)


# --------------------------------------------------------------------------- #
# Env: estado coherente de la DB como si F0.6 hubieran corrido
# --------------------------------------------------------------------------- #


def _preparar_entorno(db) -> None:
    """Estado coherente post-F0..F7 para el mini (usado como baseline PASS)."""
    from migrate.catalog import (
        bootstrap_catalogo,
        upsert_insumo,
        upsert_producto,
        upsert_proveedor,
    )

    _borrar_test(db)
    _borrar_socios_y_tipos(db)
    bootstrap_catalogo(db)
    proveedor = upsert_proveedor(db, P_PROV, url=None, ubicacion="Cali")
    tela = upsert_insumo(db, P_TELA, unidad="m", categoria_nombre="Telas")
    producto = upsert_producto(db, P_PROD, tipo="Corsetería", variantes=(P_VAR,))
    db.flush()

    # Snapshot OCT25 + compra WAC (F2+F4 del mini): stock 10, compra 2.
    tela.stock_actual = Decimal("10")
    tela.costo_promedio_actual = Decimal("100")
    db.add(CompraInsumo(
        insumo_id=tela.id,
        proveedor_id=proveedor.id,
        fecha_compra=FECHA_COMPRA,
        cantidad_comprada=Decimal("2"),
        precio_unitario_compra=Decimal("100"),
    ))

    # Socios F6 (spec FIN-2): sum == 100. FORZAMOS los % canonicos 40/30/30 en
    # cada test (un test anterior pudo alterarlos para provocar ERROR y la DB
    # real ya tiene estos socios con la migracion cargada).
    for nombre, pct in SOCIOS_ESPERADOS.items():
        socio_existente = db.query(SociosConfiguracion).filter(
            SociosConfiguracion.nombre == nombre
        ).first()
        if socio_existente is None:
            db.add(SociosConfiguracion(
                nombre=nombre, porcentaje_participacion=pct
            ))
        else:
            socio_existente.porcentaje_participacion = pct
    db.flush()

    var = next(v for v in producto.variantes if v.nombre_variante == P_VAR)
    # BOM de 'Corset': consume 2 m de tela por unidad. Igual que el apply F3
    # real, la receta se inserta con variante_id=None (BOM-1).
    db.add(BomInsumo(
        producto_id=producto.id,
        insumo_id=tela.id,
        variante_id=None,
        cantidad_requerida=Decimal("2"),
    ))
    # Cliente de la venta (el apply F5 hace upsert por nombre normalizado).
    cliente = db.query(Cliente).filter(Cliente.nombre == P_CLI).first()
    if cliente is None:
        cliente = Cliente(nombre=P_CLI)
        db.add(cliente)
        db.flush()
    # Venta historica real (VTA-1/VTA-2): costo FULL snapshot, precio tal-cual.
    venta = Venta(
        fecha=FECHA_VENTA,
        canal_venta="feria",
        descuento_porcentaje=Decimal("0"),
        total_venta=PRECIO_VENTA,  # 1 x precio tal-cual (VTA-2, sin doble desc)
        estado="completada",
        cliente_id=cliente.id,
    )
    db.add(venta)
    db.flush()
    db.add(DetalleVenta(
        venta_id=venta.id,
        producto_id=producto.id,
        variante_id=var.id,
        cantidad=Decimal("1"),
        precio_unitario_aplicado=PRECIO_VENTA,
        costo_unitario_aplicado=COSTO_VENTA,
    ))
    # Movimiento financiero F6 (equipo, socio Valqui) con fecha real.
    socio = db.query(SociosConfiguracion).filter(
        SociosConfiguracion.nombre == "Valqui"
    ).first()
    db.add(MovimientoFinanciero(
        tipo="Gasto",
        descripcion=P_MOV,
        monto=Decimal("320000"),
        fecha=FECHA_MOV,
        socio_id=socio.id if socio else None,
        estado="activo",
    ))
    db.commit()


# --------------------------------------------------------------------------- #
# Plan -> checks (reuse the future migrate.validate contract)
# --------------------------------------------------------------------------- #


def _controllers(db, mini_lib) -> dict[str, object]:
    from migrate.validate import checks_n7, plan_para_validacion

    # El mini-workbook usa una hoja CORSET cuyo bloque receta apunta al
    # producto de PRUEBA (P_PROD), no al 'Corset' real del catalogo.
    bloques_bom = {"CORSET": (P_PROD, None)}
    with LibroMigracion(mini_lib) as libro:
        plan = plan_para_validacion(libro, bloques_bom=bloques_bom)
        return {c.id: c for c in checks_n7(db, plan)}


def _no_resultado(dic, idiom: str):
    assert idiom in dic, f"check {idiom} no generado: {sorted(dic)}"
    return dic[idiom]


# --------------------------------------------------------------------------- #
# N7a: conteos por dominio
# --------------------------------------------------------------------------- #


def test_checks_generan_los_7_ids(db, mini_libro):
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    assert {k for k in res} == {
        "N7a", "N7b", "N7c", "N7d", "N7e", "N7f", "N7g"
    }


def test_n7a_conteos_coinciden_con_el_plan(db, mini_libro):
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    n7a = _no_resultado(res, "N7a")
    assert n7a.estado == "OK"
    assert "proveedores" in n7a.mensaje


def test_n7a_conteo_faltante_es_warn(db, mini_libro):
    """Plan espera un movimiento; la DB no lo tiene -> WARN (faltante con causa)."""
    from migrate.validate import plan_para_validacion

    _preparar_entorno(db)
    # Borro el movimiento test -> DB < plan.
    db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.descripcion == P_MOV
    ).delete(synchronize_session=False)
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7a").estado == "WARN"


def test_n7a_conteo_duplicado_es_error(db, mini_libro):
    """Conteo DB > plan (insumo repetido igual nombre) -> ERROR (re-run duplico)."""
    from migrate.catalog import upsert_insumo

    _preparar_entorno(db)
    dup = upsert_insumo(db, P_TELA, unidad="m", categoria_nombre="Telas")
    db.add(Insumo(
        nombre=P_TELA,  # mismo nombre -> cuenta doble en el dominio
        categoria_id=dup.categoria_id,
        unidad_medida="m",
        stock_actual=Decimal("0"),
        stock_minimo=Decimal("0"),
        costo_promedio_actual=Decimal("0"),
    ))
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7a").estado == "ERROR"


# --------------------------------------------------------------------------- #
# fix3: N7a cuenta por CLAVE NATURAL (no por nombre simple de la entidad)
# --------------------------------------------------------------------------- #


def _mini_workbook_dos_compras(path: Path) -> None:
    """Mini con DOS compras del MISMO insumo (fechas/cantidades distintas).

    Caso real del fix: Argolla 10 mm con R18 (100un) + R29 (12un) son compras
    legitimas, no un duplicado. El N7a viejo agrupaba por nombre de insumo y
    contaba 2 -> ERROR falso; el corregido agrupa por clave natural
    (insumo, fecha, cantidad, precio) -> 2 claves distintas -> OK.
    """
    _mini_workbook(path)
    wb = openpyxl.load_workbook(path)
    inv = wb["INVERSION VALQUI"]
    # R5: segunda compra del mismo insumo, fecha y cantidad distintas.
    inv.cell(row=5, column=1, value="3 mts")
    inv.cell(row=5, column=2, value=P_TELA)
    inv.cell(row=5, column=4, value=150)  # costo -> precio_unitario 50
    inv.cell(row=5, column=5, value=datetime(2025, 9, 20))
    wb.save(path)


def _mini_workbook_dos_ventas(path: Path) -> None:
    """Mini con DOS ventas del mismo producto en fechas distintas."""
    _mini_workbook(path)
    wb = openpyxl.load_workbook(path)
    ventas = wb["VENTAS"]
    ventas.append([P_PROD, P_VAR, None, None, None, None, 71250.0, COSTO_VENTA,
                   45141, None, None, None, datetime(2024, 11, 5), "vino",
                   None, P_CLI])
    wb.save(path)


def test_n7a_dos_compras_mismo_insumo_no_error(db, tmp_path):
    """fix3: 2 compras legitimas del mismo insumo (fecha/cantidad distintas)
    NO son duplicado -> N7a OK. Antes: ERROR por agrupar por nombre."""
    path = tmp_path / "mini-validate-2compras.xlsx"
    _mini_workbook_dos_compras(path)
    _preparar_entorno(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).first()
    prov = db.query(Proveedor).filter(Proveedor.nombre == P_PROV).first()
    db.add(CompraInsumo(
        insumo_id=tela.id, proveedor_id=prov.id,
        fecha_compra=FECHA_COMPRA2, cantidad_comprada=Decimal("3"),
        precio_unitario_compra=Decimal("50"),
    ))
    db.commit()
    res = _controllers(db, path)
    assert _no_resultado(res, "N7a").estado == "OK"


def test_n7a_dos_movimientos_identicos_error(db, mini_libro):
    """fix3: 2 movimientos con la MISMA clave natural -> N7a ERROR (un re-run
    de F6 habria duplicado la fila identica)."""
    _preparar_entorno(db)
    mov = db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.descripcion == P_MOV
    ).first()
    db.add(MovimientoFinanciero(
        tipo=mov.tipo, descripcion=mov.descripcion, monto=mov.monto,
        fecha=mov.fecha, socio_id=mov.socio_id, estado="activo",
    ))
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7a").estado == "ERROR"


def test_n7a_dos_ventas_fechas_distintas_no_error(db, tmp_path):
    """fix3: 2 ventas del MISMO producto en fechas distintas son legitimas
    (claves distintas) -> N7a OK. Antes: ERROR por agrupar por producto."""
    path = tmp_path / "mini-validate-2ventas.xlsx"
    _mini_workbook_dos_ventas(path)
    _preparar_entorno(db)
    producto = db.query(Producto).filter(Producto.nombre == P_PROD).first()
    var = next(v for v in producto.variantes if v.nombre_variante == P_VAR)
    cli = db.query(Cliente).filter(Cliente.nombre == P_CLI).first()
    venta2 = Venta(
        fecha=FECHA_VENTA2, canal_venta="feria", descuento_porcentaje=Decimal("0"),
        total_venta=PRECIO_VENTA, estado="completada",
        cliente_id=cli.id if cli else None,
    )
    db.add(venta2)
    db.flush()
    db.add(DetalleVenta(
        venta_id=venta2.id, producto_id=producto.id, variante_id=var.id,
        cantidad=Decimal("1"), precio_unitario_aplicado=PRECIO_VENTA,
        costo_unitario_aplicado=COSTO_VENTA,
    ))
    db.commit()
    res = _controllers(db, path)
    assert _no_resultado(res, "N7a").estado == "OK"


def test_n7a_dos_ventas_identicas_error(db, mini_libro):
    """fix3: 2 ventas EXACTAMENTE identicas (misma clave natural) -> N7a ERROR
    (un re-run de F5 habria duplicado la fila identica)."""
    _preparar_entorno(db)
    venta = db.query(Venta).filter(Venta.fecha == FECHA_VENTA).first()
    det = next(d for d in venta.detalles)
    clon = Venta(
        fecha=venta.fecha, canal_venta="feria", descuento_porcentaje=Decimal("0"),
        total_venta=venta.total_venta, estado="completada",
        cliente_id=venta.cliente_id,
    )
    db.add(clon)
    db.flush()
    db.add(DetalleVenta(
        venta_id=clon.id, producto_id=det.producto_id, variante_id=det.variante_id,
        cantidad=det.cantidad, precio_unitario_aplicado=det.precio_unitario_aplicado,
        costo_unitario_aplicado=det.costo_unitario_aplicado,
    ))
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7a").estado == "ERROR"


# --------------------------------------------------------------------------- #
# N7b: stock no negativo
# --------------------------------------------------------------------------- #


def test_n7b_stock_no_negativo_ok(db, mini_libro):
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7b").estado == "OK"


def test_n7b_stock_negativo_error(db, mini_libro):
    _preparar_entorno(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).first()
    tela.stock_actual = Decimal("-1")
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7b").estado == "ERROR"


# --------------------------------------------------------------------------- #
# N7c: finanzas (montos > 0, sumas por tipo, socios = 100)
# --------------------------------------------------------------------------- #


def test_n7c_montos_y_socios_ok(db, mini_libro):
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7c").estado == "OK"


def _mini_workbook_n7c_dup(path: Path) -> None:
    """Mini VALQUI con DOS movimientos identicos del mismo equipo.

    La fila duplicada tiene la MISMA clave natural (fecha, tipo, monto, socio,
    descripcion) -> el apply F6 la deduplica (NFR-1). N7c debe comparar contra
    el plan DEDUP (mismo criterio que el apply), no contar el duplicado 2x.
    """
    _mini_workbook(path)
    wb = openpyxl.load_workbook(path)
    inv = wb["INVERSION VALQUI"]
    # R5: segundo movimiento identico a R4 (P_MOV 320000 @ FECHA_MOV).
    inv.cell(row=5, column=1, value=1)
    inv.cell(row=5, column=2, value=P_MOV)
    inv.cell(row=5, column=4, value=320000)
    inv.cell(row=5, column=5, value=datetime(2025, 8, 1))
    wb.save(path)


def test_n7c_plan_movimiento_duplicado_cuenta_una_vez(db, tmp_path):
    """P3 fix: movimientos del plan con clave natural duplicada se cuentan UNA
    vez (el apply F6 deduplica por clave, NFR-1). Sin el fix la suma del plan
    (640000) difiere de la DB (320000) -> ERROR."""
    path = tmp_path / "mini-validate-dup.xlsx"
    _mini_workbook_n7c_dup(path)
    _preparar_entorno(db)
    res = _controllers(db, path)
    n7c = _no_resultado(res, "N7c")
    assert n7c.estado == "OK"
    assert "duplicada" in n7c.mensaje  # desviacion conocida documentada


def test_n7c_monto_cero_es_error(db, mini_libro):
    _preparar_entorno(db)
    mov = db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.descripcion == P_MOV
    ).first()
    mov.monto = Decimal("0")
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7c").estado == "ERROR"


def test_n7c_socios_menos_cien_error(db, mini_libro):
    _preparar_entorno(db)
    marg = db.query(SociosConfiguracion).filter(
        SociosConfiguracion.nombre == "Margarita"
    ).first()
    marg.porcentaje_participacion = Decimal("40")
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7c").estado == "ERROR"


# --------------------------------------------------------------------------- #
# N7d: cuadre de stock (snapshot + compras - explosiones de ventas)
# --------------------------------------------------------------------------- #


def test_n7d_cuadre_exacto(db, mini_libro):
    """Snapshot 10 + compra 2 - explosion de 1 venta (Corset: 2 m) = 10."""
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7d").estado == "OK"


# Fecha del corte fisico del snapshot OCT25 (design: stock OCT25 = 2025-10-25).
FECHA_CORTE_OCT25 = datetime(2025, 10, 25, tzinfo=timezone.utc)
FECHA_POST_CORTE = datetime(2025, 11, 1, tzinfo=timezone.utc)


def test_n7d_compra_del_corte_no_se_suma(db, mini_libro):
    """P2 fix: la compra que constituye el corte fisico (fecha == OCT25) no se
    suma al cuadre: el snapshot YA la contiene (Ref 100: 39 + 0 - consumos)."""
    _preparar_entorno(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).first()
    db.add(CompraInsumo(
        insumo_id=tela.id,
        proveedor_id=None,
        fecha_compra=FECHA_CORTE_OCT25,
        cantidad_comprada=Decimal("10"),
        precio_unitario_compra=Decimal("100"),
    ))
    db.commit()
    res = _controllers(db, mini_libro)
    # esperado = snapshot 10 + compra pre-corte 2 + 0 (corte excluida) - 2 = 10.
    assert _no_resultado(res, "N7d").estado == "OK"


def test_n7d_compra_post_corte_si_suma(db, mini_libro):
    """P2 fix: compras POSTERIORES al corte (fecha > OCT25) SI suman al stock
    del corte (son compras nuevas despues del inventario)."""
    _preparar_entorno(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).first()
    db.add(CompraInsumo(
        insumo_id=tela.id,
        proveedor_id=None,
        fecha_compra=FECHA_POST_CORTE,
        cantidad_comprada=Decimal("3"),
        precio_unitario_compra=Decimal("100"),
    ))
    tela.stock_actual = Decimal("13")  # 10 + 3 post-corte - 2 consumidos
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7d").estado == "OK"


def test_n7d_cuadre_divergente_warn(db, mini_libro):
    _preparar_entorno(db)
    tela = db.query(Insumo).filter(Insumo.nombre == P_TELA).first()
    tela.stock_actual = Decimal("9")  # 10 esperado -> divergencia con causa
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7d").estado == "WARN"


def test_n7d_snapshot_insumo_ausente_error(db, tmp_path):
    """Snapshot pide un insumo que no esta en el catalogo -> ERROR precondicion."""
    path = tmp_path / "mini-validate-extra.xlsx"
    _mini_workbook(path, extra_oct25=(P_EXTRA, "2 mts"))
    _preparar_entorno(db)
    res = _controllers(db, path)
    assert _no_resultado(res, "N7d").estado == "ERROR"


# --------------------------------------------------------------------------- #
# N7e: precios de venta (costo historico + total sin doble descuento)
# --------------------------------------------------------------------------- #


def test_n7e_precios_ok(db, mini_libro):
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7e").estado == "OK"


def test_n7e_costo_alterado_error(db, mini_libro):
    _preparar_entorno(db)
    det = (
        db.query(DetalleVenta)
        .join(Venta, DetalleVenta.venta_id == Venta.id)
        .filter(Venta.fecha == FECHA_VENTA)
        .first()
    )
    det.costo_unitario_aplicado = Decimal("199")
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7e").estado == "ERROR"


def test_n7e_total_con_descuento_doble_error(db, mini_libro):
    _preparar_entorno(db)
    venta = db.query(Venta).filter(Venta.fecha == FECHA_VENTA).first()
    venta.descuento_porcentaje = Decimal("25")  # precio ya descontado (VTA-2)
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7e").estado == "ERROR"


# --------------------------------------------------------------------------- #
# N7f: fechas reales (nunca now())
# --------------------------------------------------------------------------- #


def test_n7f_fechas_reales_ok(db, mini_libro):
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7f").estado == "OK"


def test_n7f_compra_con_fecha_now_error(db, mini_libro):
    """Compra del insumo de PRUEBA con fecha now() -> ERROR. Se filtra por el
    insumo test: con la migracion cargada, `first()` sin filtro tomarla una
    compra real fuera de alcance y la corromperia."""
    _preparar_entorno(db)
    compra = (
        db.query(CompraInsumo)
        .join(Insumo, CompraInsumo.insumo_id == Insumo.id)
        .filter(Insumo.nombre == P_TELA)
        .first()
    )
    compra.fecha_compra = datetime.now(timezone.utc)
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7f").estado == "ERROR"


# --------------------------------------------------------------------------- #
# N7g: idempotencia (claves naturales duplicadas)
# --------------------------------------------------------------------------- #


def test_n7g_sin_duplicados_ok(db, mini_libro):
    _preparar_entorno(db)
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7g").estado == "OK"


def test_n7g_compra_duplicada_error(db, mini_libro):
    """Compra del insumo de PRUEBA duplicada (misma clave natural F2) -> ERROR.
    Se filtra por el insumo test: con la migracion cargada, `first()` sin
    filtro tomarla una compra real fuera del alcance del plan mini."""
    _preparar_entorno(db)
    compra = (
        db.query(CompraInsumo)
        .join(Insumo, CompraInsumo.insumo_id == Insumo.id)
        .filter(Insumo.nombre == P_TELA)
        .first()
    )
    db.add(CompraInsumo(
        insumo_id=compra.insumo_id,
        proveedor_id=compra.proveedor_id,
        fecha_compra=compra.fecha_compra,
        cantidad_comprada=compra.cantidad_comprada,
        precio_unitario_compra=compra.precio_unitario_compra,
    ))
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7g").estado == "ERROR"


def test_n7g_movimiento_duplicado_error(db, mini_libro):
    _preparar_entorno(db)
    mov = db.query(MovimientoFinanciero).filter(
        MovimientoFinanciero.descripcion == P_MOV
    ).first()
    db.add(MovimientoFinanciero(
        tipo=mov.tipo,
        descripcion=mov.descripcion,
        monto=mov.monto,
        fecha=mov.fecha,
        socio_id=mov.socio_id,
        estado="activo",
    ))
    db.commit()
    res = _controllers(db, mini_libro)
    assert _no_resultado(res, "N7g").estado == "ERROR"


# --------------------------------------------------------------------------- #
# Fase runner + dry-run real (NFR-2)
# --------------------------------------------------------------------------- #


def test_f7_registrada_en_runner():
    from migrate import FASE_RUNNERS, FASES, FASES_IMPLEMENTADAS

    assert any(f.id == "F7" for f in FASES)
    assert "F7" in FASE_RUNNERS
    assert "F7" in FASES_IMPLEMENTADAS


def test_cargar_validate_dry_run_real_no_escribe():
    """F7 dry-run sobre ARPIA.xlsx real: valida sin escribir nada (NFR-2)."""
    from migrate.validate import cargar_validate

    if not REAL_XLSX.exists():
        pytest.skip("ARPIA.xlsx no disponible")

    from app.models import Insumo, MovimientoFinanciero, Producto

    db = SessionLocal()
    try:
        antes = {
            "insumos": db.query(Insumo).count(),
            "movimientos": db.query(MovimientoFinanciero).count(),
            "productos": db.query(Producto).count(),
        }
        ctx = MigrationContext.para_fase(
            FaseOptions(source=REAL_XLSX, modo="dry-run"), "F7"
        )
        cargar_validate(ctx)
        despues = {
            "insumos": db.query(Insumo).count(),
            "movimientos": db.query(MovimientoFinanciero).count(),
            "productos": db.query(Producto).count(),
        }
        assert antes == despues  # NFR-2: 0 filas escritas
        assert not ctx.report.tenga_errores  # dry-run: sin ERROR (validacion sola)
    finally:
        db.close()
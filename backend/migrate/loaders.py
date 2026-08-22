"""Excel loaders: bounded, per-sheet reading of ARPIA.xlsx (openpyxl 3.1.5).

Design (EXM-1): iterate ONLY the real used range per sheet. Several sheets have
inflated max_row (BOM matrices have junk blocks and ghost sub-tables). Each
sheet declares bounds via ``SHEET_BOUNDS`` and known traps:

- M37 in VENTAS is a loose outlier cell -> reported with sheet/row/cell;
- CAMISETAS INV rows R10-13 are junk (real data ends at R9);
- #DIV/0! values in INVERSION VALQUI N70-N78 -> error value, row excluded;
- ghost sub-tabs (TANGA columns in BOM sheets) are not real data rows;
- CAMISETAS INV header is misaligned: col D=tipo, E=origen, F=color real data
  (the physical headers say otherwise) -> per-sheet column mapping.

The loader returns per-sheet: real rows (dict col_letter -> value), counts of
discarded/omitted rows, and reported issues (sheet/row/cell) so the phase
report can log them (N7g).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from migrate.report import Report

# Per-sheet bounds: (first_data_row, last_data_row). Header rows precede the
# first data row unless noted. Values verified against ARPIA.xlsx (2026-08-08).
SHEET_BOUNDS: dict[str, tuple[int, int]] = {
    "CAMISETAS": (3, 5),  # header R2 (Referencia..); junk R6-13
    "ARPIA": (2, 9),  # header R1; matrix R2-9; junk R10-14
    "INVERSION VALQUI": (3, 137),  # header R2; sub-tab herrajes J-N right
    "INVERSION MARGARA": (3, 70),  # header R2; multi-tab right H-J
    "IDEAS": (0, 0),  # no data rows (single URL cell only)
    "STICKERS": (2, 33),  # header R1
    "CAMISETAS INV": (2, 9),  # header R1 (misaligned cols); junk R10-17
    "Braleth dise\u00f1o 1": (3, 24),  # header R2
    "Noche y Dia CACHETERO": (3, 24),
    "Noche y Dia": (3, 30),
    "CORSET": (3, 33),
    "CORSET DOBLE CARA": (3, 33),
    "CORSET ARTEMISIA": (3, 33),
    "FALDA EMILY": (3, 33),
    "Corset Hypatia": (3, 33),
    "BUSTIER": (3, 33),
    "BLUSAS": (3, 45),
    "TOTEBAG": (3, 44),
    # Recalculated 16-sheet workbook (2026-08): BOM recipe sheets with the
    # standard layout (header R2: Producto/Ancho/Alto/cantidad Cms/valor metro/
    # valor total). Real data ends at the last bound row; junk/#VALUE!/totals
    # rows live after it and are never read.
    "SET AELO": (3, 21),
    "SET OCIPETE": (3, 18),
    "Corset Garras": (3, 10),
    "DESCUENTOS": (3, 21),  # header R2; block header repeats at R13
    "INVENTARIO OCT25": (9, 29),  # header R8; totals row R37
    "CAJAS": (4, 13),  # headers R2-3; totals R13-14
    "VENTAS": (2, 17),  # header R1; 16 real rows; junk/zeros >17
    "GASTOS ARPIA": (5, 8),  # header R4; block headers at R16+...
}

# Misaligned header: physical cols vs real semantics. CAMISETAS INV header says
# E=Fecha, F=Provedor but real data rows put Tipo/Origen/Color in D/E/F.
COL_MAP_CAMISETAS_INV = {
    "A": "Cantidad",
    "B": "Talla",
    "C": "Costo",
    "D": "Tipo",
    "E": "Origen",
    "F": "Color",
}

# Per-sheet column maps (EXM-1 error scenario): sheets whose physical header is
# desalineado declare their real semantics; ``leer_hoja`` applies the mapping to
# every row (key = semantic name) and reports the mismatch against the expected
# header. Sheets without an entry keep the physical column-letter keys.
COL_MAPS: dict[str, dict[str, str]] = {
    "CAMISETAS INV": COL_MAP_CAMISETAS_INV,
}

# Known junk cells to report explicitly (sheet name -> (col letter, row)).
JUMP_CELLS: dict[str, list[tuple[str, int]]] = {
    "VENTAS": [("M", 37)],
}

# '#'-prefixed Excel error values must exclude a row (never inferred, EXM-2).
ERROR_VALUE_PREFIX = "#"


class HojaInexistenteError(Exception):
    """Raised when the requested sheet is not present in the source workbook."""

    def __init__(self, hoja: str, source: Path):
        self.hoja = hoja
        self.source = source
        super().__init__(f"La hoja {hoja!r} no existe en {source}")


@dataclass
class LecturaHoja:
    """Result of reading one sheet: real rows + counts + reported issues."""

    hoja: str
    filas: list[dict[str, object]] = field(default_factory=list)  # one dict per real row
    descartadas: int = 0  # empty rows inside the bound (not counted as data)
    omitidas: int = 0  # rows excluded due to an ERROR_VALUE (e.g. #DIV/0!)
    issues: list[tuple[str, int, str, str]] = field(default_factory=list)  # (col,row,celda,msg)


@dataclass
class LibroMigracion:
    """Bounded Excel source wrapper (openpyxl read_only + data_only)."""

    ruta: Path
    _libro: object = field(default=None, repr=False)

    def abrir(self) -> None:
        self._libro = load_workbook(self.ruta, data_only=True, read_only=True)

    @property
    def hojas(self) -> list[str]:
        if self._libro is None:
            self.abrir()
        return list(self._libro.sheetnames)

    def obtener_worksheet(self, hoja: str) -> Worksheet:
        if self._libro is None:
            self.abrir()
        if hoja not in self._libro.sheetnames:
            raise HojaInexistenteError(hoja, self.ruta)
        return self._libro[hoja]

    def leer_hoja(self, hoja: str, report: Report | None = None) -> LecturaHoja:
        """Read a single sheet within its declared bounds (EXM-1)."""
        ws = self.obtener_worksheet(hoja)
        resultado = LecturaHoja(hoja=hoja)

        if hoja not in SHEET_BOUNDS:
            if report is not None:
                report.warn(hoja, None, None, "hoja sin rango registrado; ignorada (0 filas)")
            return resultado

        first_row, last_row = SHEET_BOUNDS[hoja]
        if first_row > last_row or first_row == 0:
            # e.g. IDEAS: no data rows by design.
            return resultado

        # EXM-1 error scenario: hoja con encabezado desalineado -> se aplica el
        # mapeo de columnas propio, comprobado contra el encabezado esperado.
        mapeo = COL_MAPS.get(hoja)
        if mapeo is not None and report is not None:
            header = ws.cell(row=1, column=column_index_from_string("D")).value
            esperado = mapeo.get("D")
            if header is None or str(header).strip() != esperado:
                report.warn(
                    hoja,
                    1,
                    "D1",
                    f"encabezado desalineado: fila real D/E/F = "
                    f"Tipo/Origen/Color pero header dice {header!r}; "
                    f"aplicado mapeo de columnas propio (EXM-1)",
                )

        for fila_idx, row in enumerate(
            ws.iter_rows(min_row=first_row, max_row=last_row), start=first_row
        ):
            celdas: dict[str, object] = {}
            error_celda: str | None = None
            for celda in row:
                v = celda.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith(ERROR_VALUE_PREFIX):
                    error_celda = celda.coordinate
                    break
                celdas[celda.column_letter] = v

            if error_celda is not None:
                # #DIV/0! (or similar): exclude the row, never infer (EXM-2).
                resultado.omitidas += 1
                resultado.issues.append(
                    (
                        None,
                        fila_idx,
                        error_celda,
                        f"valor de error en {error_celda} -> fila excluida (no se infiere)",
                    )
                )
                continue

            if not celdas:
                # Fully empty row inside the bound -> discarded (not data).
                resultado.descartadas += 1
                continue

            if mapeo is not None:
                celdas = {mapeo.get(k, k): v for k, v in celdas.items()}
            resultado.filas.append(celdas)

        # Explicit junk cells (M37 suelta in VENTAS): report, never data.
        for col_letter, fila_junk in JUMP_CELLS.get(hoja, []):
            cell = ws.cell(
                row=fila_junk,
                column=column_index_from_string(col_letter),
            )
            if cell.value is not None:
                coord = f"{col_letter}{fila_junk}"
                if report is not None:
                    report.warn(
                        hoja,
                        fila_junk,
                        coord,
                        f"celda suelta {cell.value!r} ignorada (no es fila de datos)",
                    )
                resultado.issues.append((col_letter, fila_junk, coord, "celda suelta ignorada"))
        return resultado

    def cerrar(self) -> None:
        if self._libro is not None:
            self._libro.close()
            self._libro = None

    def __enter__(self) -> LibroMigracion:
        self.abrir()
        return self

    def __exit__(self, *exc) -> None:
        self.cerrar()
